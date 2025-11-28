"""
OHTC 專案管理儀表板
================================
功能：
- 讀取 Excel 排程表
- 視覺化甘特圖、進度圖
- 追蹤延遲項目
- 編輯並匯出 Excel
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import warnings
warnings.filterwarnings('ignore')

# 頁面設定
st.set_page_config(
    page_title="OHTC 專案管理儀表板",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 自訂 CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 1rem;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        text-align: center;
    }
    .status-done { background-color: #28a745; color: white; padding: 3px 8px; border-radius: 4px; }
    .status-going { background-color: #ffc107; color: black; padding: 3px 8px; border-radius: 4px; }
    .status-delay { background-color: #dc3545; color: white; padding: 3px 8px; border-radius: 4px; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] {
        background-color: #f0f2f6;
        border-radius: 4px 4px 0 0;
        padding: 10px 20px;
    }
</style>
""", unsafe_allow_html=True)


def load_excel_data(uploaded_file):
    """載入 Excel 檔案並解析各工作表"""
    try:
        # 讀取軟體時程表
        df_software = pd.read_excel(uploaded_file, sheet_name='軟體時程', header=None)
        
        # 提取專案資訊
        project_info = {
            'project_code': str(df_software.iloc[2, 2]) if pd.notna(df_software.iloc[2, 2]) else '',
            'project_name': str(df_software.iloc[3, 2]) if pd.notna(df_software.iloc[3, 2]) else '',
            'project_lead': str(df_software.iloc[4, 2]) if pd.notna(df_software.iloc[4, 2]) else '',
            'start_date': df_software.iloc[3, 9] if pd.notna(df_software.iloc[3, 9]) else None,
        }
        
        # 解析任務資料（從第7行開始）
        tasks = []

        # 安全轉換函數
        def safe_float(val, default=0):
            try:
                if pd.isna(val):
                    return default
                if isinstance(val, str) and not val.replace('.', '', 1).replace('-', '', 1).isdigit():
                    return default
                return float(val)
            except (ValueError, TypeError):
                return default

        def safe_int(val, default=0):
            try:
                if pd.isna(val):
                    return default
                if isinstance(val, str) and not val.replace('-', '', 1).isdigit():
                    return default
                return int(float(val))
            except (ValueError, TypeError):
                return default

        for i in range(6, len(df_software)):
            row = df_software.iloc[i]
            task_name = row[0]

            if pd.notna(task_name) and str(task_name).strip():
                # 跳過標題行
                if isinstance(row[4], str) and ('百分比' in str(row[4]) or '完成' in str(row[4])):
                    continue

                task = {
                    'id': len(tasks) + 1,
                    'task': str(task_name).strip(),
                    'owner': str(row[2]) if pd.notna(row[2]) else '',
                    'progress_pct': safe_float(row[4]),
                    'target_pct': safe_float(row[5]),
                    'remaining_days': safe_int(row[6]),
                    'status': str(row[7]) if pd.notna(row[7]) else '',
                    'plan_start': row[8] if pd.notna(row[8]) else None,
                    'plan_end': row[9] if pd.notna(row[9]) else None,
                    'plan_days': safe_int(row[10]),
                    'actual_start': row[11] if pd.notna(row[11]) else None,
                    'actual_end': row[12] if pd.notna(row[12]) else None,
                    'actual_days': safe_int(row[13]),
                    'variance_days': safe_int(row[14]),
                    'notes': str(row[19]) if pd.notna(row[19]) else '',
                }
                tasks.append(task)
        
        df_tasks = pd.DataFrame(tasks)
        
        # 讀取系統時程
        df_system = pd.read_excel(uploaded_file, sheet_name='系統時程_C', header=None)
        system_items = []
        for i in range(5, len(df_system)):
            row = df_system.iloc[i]
            if pd.notna(row[0]):
                # 跳過標題行
                if isinstance(row[0], str) and '區域' in str(row[0]) and i == 5:
                    continue

                item = {
                    'item': str(row[0]).strip(),
                    'target_date': row[1] if pd.notna(row[1]) else None,
                    'completion_pct': safe_float(row[2]),
                    'notes': str(row[3]) if pd.notna(row[3]) else '',
                }
                system_items.append(item)
        df_system_tasks = pd.DataFrame(system_items)
        
        return {
            'project_info': project_info,
            'tasks': df_tasks,
            'system_tasks': df_system_tasks,
            'raw_software': df_software,
        }
    except Exception as e:
        st.error(f"載入檔案錯誤: {str(e)}")
        return None


def create_gantt_chart(df_tasks):
    """建立甘特圖"""
    # 過濾有效資料
    gantt_data = df_tasks[df_tasks['plan_start'].notna() & df_tasks['plan_end'].notna()].copy()
    
    if gantt_data.empty:
        return None
    
    # 確保日期格式
    gantt_data['plan_start'] = pd.to_datetime(gantt_data['plan_start'])
    gantt_data['plan_end'] = pd.to_datetime(gantt_data['plan_end'])
    
    # 狀態顏色對應
    color_map = {
        'Done': '#28a745',
        'Going': '#ffc107', 
        'Delay': '#dc3545',
        '': '#6c757d'
    }
    
    gantt_data['color'] = gantt_data['status'].map(lambda x: color_map.get(x, '#6c757d'))
    
    fig = px.timeline(
        gantt_data,
        x_start='plan_start',
        x_end='plan_end',
        y='task',
        color='status',
        color_discrete_map=color_map,
        hover_data=['owner', 'plan_days', 'variance_days'],
        title='📅 專案甘特圖 (計劃時程)'
    )
    
    fig.update_layout(
        height=max(400, len(gantt_data) * 25),
        xaxis_title='日期',
        yaxis_title='',
        yaxis={'categoryorder': 'total ascending'},
        showlegend=True,
        legend_title='狀態',
    )
    
    # 加入今日線
    today = datetime.now()
    fig.add_vline(x=today, line_dash="dash", line_color="red", annotation_text="今日")
    
    return fig


def create_status_chart(df_tasks):
    """建立狀態圓餅圖"""
    status_counts = df_tasks['status'].value_counts()
    
    colors = {
        'Done': '#28a745',
        'Going': '#ffc107',
        'Delay': '#dc3545',
    }
    
    fig = px.pie(
        values=status_counts.values,
        names=status_counts.index,
        title='📊 任務狀態分佈',
        color=status_counts.index,
        color_discrete_map=colors,
        hole=0.4
    )
    
    fig.update_traces(textposition='inside', textinfo='value+percent')
    fig.update_layout(height=400)
    
    return fig


def create_owner_chart(df_tasks):
    """建立負責單位工作量圖"""
    owner_counts = df_tasks.groupby('owner').agg({
        'task': 'count',
        'status': lambda x: (x == 'Done').sum()
    }).reset_index()
    owner_counts.columns = ['owner', 'total', 'done']
    owner_counts = owner_counts[owner_counts['owner'] != '']
    owner_counts['pending'] = owner_counts['total'] - owner_counts['done']
    
    fig = go.Figure()
    fig.add_trace(go.Bar(name='已完成', x=owner_counts['owner'], y=owner_counts['done'], marker_color='#28a745'))
    fig.add_trace(go.Bar(name='進行中', x=owner_counts['owner'], y=owner_counts['pending'], marker_color='#ffc107'))
    
    fig.update_layout(
        barmode='stack',
        title='👥 各負責單位工作量',
        xaxis_title='負責單位',
        yaxis_title='任務數量',
        height=400
    )
    
    return fig


def create_progress_gauge(completed, total, title):
    """建立進度儀表板"""
    pct = (completed / total * 100) if total > 0 else 0
    
    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=pct,
        domain={'x': [0, 1], 'y': [0, 1]},
        title={'text': title, 'font': {'size': 16}},
        number={'suffix': '%', 'font': {'size': 30}},
        gauge={
            'axis': {'range': [0, 100], 'tickwidth': 1},
            'bar': {'color': "#28a745" if pct >= 70 else "#ffc107" if pct >= 40 else "#dc3545"},
            'steps': [
                {'range': [0, 40], 'color': '#ffebee'},
                {'range': [40, 70], 'color': '#fff8e1'},
                {'range': [70, 100], 'color': '#e8f5e9'}
            ],
            'threshold': {
                'line': {'color': "red", 'width': 2},
                'thickness': 0.75,
                'value': 90
            }
        }
    ))
    
    fig.update_layout(height=250, margin=dict(l=20, r=20, t=40, b=20))
    return fig


def export_to_excel(data, original_file):
    """匯出更新後的資料到 Excel"""
    output = io.BytesIO()
    
    # 複製原始檔案
    original_file.seek(0)
    wb = load_workbook(original_file)
    
    # 更新軟體時程表
    ws = wb['軟體時程']
    df_tasks = data['tasks']
    
    # 從第7行開始更新
    for idx, task in df_tasks.iterrows():
        row_num = idx + 7  # Excel 行號
        # 更新狀態欄
        ws.cell(row=row_num, column=8, value=task['status'])
        # 可以根據需要更新其他欄位
    
    wb.save(output)
    output.seek(0)
    return output


def main():
    st.markdown('<h1 class="main-header">🏭 OHTC 專案管理儀表板</h1>', unsafe_allow_html=True)
    
    # 側邊欄
    with st.sidebar:
        st.header("📁 檔案上傳")
        uploaded_file = st.file_uploader(
            "上傳專案排程表 (.xlsx)",
            type=['xlsx', 'xls'],
            help="請上傳 OHTC 安裝排程表 Excel 檔案"
        )
        
        st.divider()
        
        if uploaded_file:
            st.success("✅ 檔案已載入")
            st.info(f"📄 {uploaded_file.name}")
    
    if uploaded_file is None:
        # 顯示說明
        st.info("👆 請先上傳專案排程表 Excel 檔案")
        
        st.markdown("""
        ### 📌 功能說明
        
        此工具可以：
        1. **視覺化呈現** - 甘特圖、進度圖、狀態分佈
        2. **追蹤進度** - 即時顯示延遲項目和待辦事項
        3. **團隊協作** - 各負責單位工作量一目了然
        4. **匯出報表** - 保持原 Excel 格式匯出
        
        ### 📋 支援的 Excel 格式
        - 軟體時程表（甘特圖資料）
        - 系統時程表（區域進度）
        - 工程進度確認表
        """)
        return
    
    # 載入資料
    data = load_excel_data(uploaded_file)
    
    if data is None:
        return
    
    project_info = data['project_info']
    df_tasks = data['tasks']
    df_system = data['system_tasks']
    
    # 專案資訊卡
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("📋 專案工令", project_info['project_code'])
    with col2:
        st.metric("📌 專案名稱", project_info['project_name'][:20] + "..." if len(project_info['project_name']) > 20 else project_info['project_name'])
    with col3:
        st.metric("👤 專案負責", project_info['project_lead'])
    with col4:
        if project_info['start_date']:
            st.metric("📅 開始日期", pd.to_datetime(project_info['start_date']).strftime('%Y-%m-%d'))
    
    st.divider()
    
    # 關鍵指標
    total_tasks = len(df_tasks)
    done_tasks = len(df_tasks[df_tasks['status'] == 'Done'])
    delay_tasks = len(df_tasks[df_tasks['status'] == 'Delay'])
    going_tasks = len(df_tasks[df_tasks['status'] == 'Going'])
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        fig = create_progress_gauge(done_tasks, total_tasks, "整體完成率")
        st.plotly_chart(fig, use_container_width=True)
    with col2:
        st.metric("📝 總任務數", total_tasks)
        st.metric("✅ 已完成", done_tasks, delta=f"{done_tasks/total_tasks*100:.1f}%" if total_tasks > 0 else "0%")
    with col3:
        st.metric("🔄 進行中", going_tasks)
        st.metric("⚠️ 延遲中", delay_tasks, delta=f"-{delay_tasks}" if delay_tasks > 0 else None, delta_color="inverse")
    with col4:
        avg_variance = df_tasks[df_tasks['variance_days'] != 0]['variance_days'].mean()
        st.metric("📊 平均誤差天數", f"{avg_variance:.1f}" if pd.notna(avg_variance) else "N/A")
    
    st.divider()
    
    # 主要內容區 - 標籤頁
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📅 甘特圖", 
        "📊 統計圖表", 
        "⚠️ 延遲追蹤",
        "📋 任務清單",
        "⬇️ 匯出"
    ])
    
    with tab1:
        st.subheader("📅 專案甘特圖")
        gantt_fig = create_gantt_chart(df_tasks)
        if gantt_fig:
            st.plotly_chart(gantt_fig, use_container_width=True)
        else:
            st.warning("沒有足夠的資料來建立甘特圖")
    
    with tab2:
        col1, col2 = st.columns(2)
        with col1:
            status_fig = create_status_chart(df_tasks)
            st.plotly_chart(status_fig, use_container_width=True)
        with col2:
            owner_fig = create_owner_chart(df_tasks)
            st.plotly_chart(owner_fig, use_container_width=True)
        
        # 系統時程進度
        st.subheader("🔧 系統時程進度 (按區域)")
        if not df_system.empty:
            # 篩選區域項目
            area_items = df_system[df_system['item'].str.contains('區域', na=False)]
            if not area_items.empty:
                fig = px.bar(
                    area_items,
                    x='item',
                    y='completion_pct',
                    title='各區域完成進度',
                    color='completion_pct',
                    color_continuous_scale='RdYlGn'
                )
                fig.update_layout(yaxis_range=[0, 1], yaxis_tickformat='.0%')
                st.plotly_chart(fig, use_container_width=True)
    
    with tab3:
        st.subheader("⚠️ 延遲項目追蹤")
        
        delay_df = df_tasks[df_tasks['status'] == 'Delay']
        
        if delay_df.empty:
            st.success("🎉 太棒了！目前沒有延遲的項目！")
        else:
            st.error(f"⚠️ 共有 {len(delay_df)} 個延遲項目需要關注")
            
            for _, task in delay_df.iterrows():
                with st.expander(f"🔴 {task['task']}", expanded=True):
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.write(f"**負責單位:** {task['owner']}")
                    with col2:
                        st.write(f"**計劃完成:** {pd.to_datetime(task['plan_end']).strftime('%Y-%m-%d') if pd.notna(task['plan_end']) else 'N/A'}")
                    with col3:
                        st.write(f"**誤差天數:** {task['variance_days']} 天")
        
        st.divider()
        
        # 即將到期項目
        st.subheader("⏰ 即將到期項目 (7天內)")
        today = datetime.now()
        upcoming = df_tasks[
            (df_tasks['status'] == 'Going') & 
            (pd.to_datetime(df_tasks['plan_end']) <= today + timedelta(days=7)) &
            (pd.to_datetime(df_tasks['plan_end']) >= today)
        ]
        
        if upcoming.empty:
            st.info("近期沒有即將到期的項目")
        else:
            for _, task in upcoming.iterrows():
                days_left = (pd.to_datetime(task['plan_end']) - today).days
                st.warning(f"⏰ **{task['task']}** - 剩餘 {days_left} 天 (負責: {task['owner']})")
    
    with tab4:
        st.subheader("📋 完整任務清單")
        
        # 篩選器
        col1, col2, col3 = st.columns(3)
        with col1:
            status_filter = st.multiselect(
                "篩選狀態",
                options=['Done', 'Going', 'Delay', ''],
                default=['Done', 'Going', 'Delay']
            )
        with col2:
            owners = df_tasks['owner'].unique().tolist()
            owner_filter = st.multiselect(
                "篩選負責單位",
                options=owners,
                default=[]
            )
        with col3:
            search = st.text_input("🔍 搜尋任務名稱")
        
        # 套用篩選
        filtered_df = df_tasks[df_tasks['status'].isin(status_filter)]
        if owner_filter:
            filtered_df = filtered_df[filtered_df['owner'].isin(owner_filter)]
        if search:
            filtered_df = filtered_df[filtered_df['task'].str.contains(search, case=False, na=False)]
        
        # 顯示表格
        display_cols = ['task', 'owner', 'status', 'plan_start', 'plan_end', 'plan_days', 'actual_start', 'actual_end', 'variance_days']
        display_df = filtered_df[display_cols].copy()
        display_df.columns = ['任務', '負責單位', '狀態', '計劃開始', '計劃完成', '計劃天數', '實際開始', '實際完成', '誤差天數']
        
        st.dataframe(
            display_df,
            use_container_width=True,
            height=500,
            column_config={
                "狀態": st.column_config.SelectboxColumn(
                    options=["Done", "Going", "Delay"],
                ),
                "計劃開始": st.column_config.DateColumn(format="YYYY-MM-DD"),
                "計劃完成": st.column_config.DateColumn(format="YYYY-MM-DD"),
                "實際開始": st.column_config.DateColumn(format="YYYY-MM-DD"),
                "實際完成": st.column_config.DateColumn(format="YYYY-MM-DD"),
            }
        )
        
        st.caption(f"顯示 {len(filtered_df)} / {len(df_tasks)} 筆資料")
    
    with tab5:
        st.subheader("⬇️ 匯出報表")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### 📄 匯出 Excel")
            st.write("保持原始格式，匯出更新後的排程表")
            
            if st.button("🔄 生成 Excel 檔案", type="primary"):
                try:
                    excel_output = export_to_excel(data, uploaded_file)
                    st.download_button(
                        label="⬇️ 下載 Excel",
                        data=excel_output,
                        file_name=f"OHTC_排程表_更新_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"匯出失敗: {str(e)}")
        
        with col2:
            st.markdown("### 📊 匯出 CSV")
            st.write("匯出任務清單為 CSV 格式")
            
            csv = df_tasks.to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                label="⬇️ 下載 CSV",
                data=csv,
                file_name=f"OHTC_任務清單_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv"
            )


if __name__ == "__main__":
    main()
