"""
OHTC 專案管理儀表板 v2.0
================================
新增功能：
- 週報/月報自動生成
- 進度趨勢分析
- 風險評估矩陣
- 資源負載分析
- 里程碑追蹤
- 任務編輯與儲存
- 多工作表完整解析
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import io
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

# 導入通知模組
try:
    from notifications import NotificationConfig, ProjectNotifier
    NOTIFICATIONS_AVAILABLE = True
except ImportError:
    NOTIFICATIONS_AVAILABLE = False

# 導入模板生成器
try:
    from template_generator import ScheduleTemplateGenerator
    TEMPLATE_GENERATOR_AVAILABLE = True
except ImportError:
    TEMPLATE_GENERATOR_AVAILABLE = False

# ============================================================
# 頁面設定
# ============================================================
st.set_page_config(
    page_title="OHTC 專案管理儀表板 v2.0",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="auto"  # 手機端自動收合側邊欄
)

# 自訂 CSS
st.markdown("""
<style>
    /* 基礎樣式 */
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        background: linear-gradient(90deg, #1f77b4, #9467bd);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        margin-bottom: 1rem;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        text-align: center;
    }
    .risk-high { background-color: #dc3545; color: white; padding: 5px 10px; border-radius: 4px; }
    .risk-medium { background-color: #ffc107; color: black; padding: 5px 10px; border-radius: 4px; }
    .risk-low { background-color: #28a745; color: white; padding: 5px 10px; border-radius: 4px; }
    .milestone-done { border-left: 4px solid #28a745; }
    .milestone-pending { border-left: 4px solid #ffc107; }
    .report-section {
        background: #f8f9fa;
        border-radius: 8px;
        padding: 20px;
        margin: 10px 0;
    }
    div[data-testid="stExpander"] details summary p {
        font-size: 1.1rem;
        font-weight: 600;
    }

    /* 響應式設計 - 手機端優化 */
    @media only screen and (max-width: 768px) {
        /* 主標題縮小 */
        .main-header {
            font-size: 1.5rem;
        }

        /* 全局文字換行 */
        body, div, p, span, li, td, th {
            word-wrap: break-word !important;
            word-break: break-word !important;
            overflow-wrap: break-word !important;
        }

        /* Streamlit 容器優化 */
        .stApp {
            max-width: 100vw;
            overflow-x: hidden;
        }

        /* 表格優化 */
        div[data-testid="stDataFrame"] {
            overflow-x: auto !important;
            max-width: 100vw !important;
        }

        /* Plotly 圖表優化 */
        .js-plotly-plot {
            overflow-x: auto !important;
            max-width: 100vw !important;
        }

        /* 圖表容器 */
        div[data-testid="stPlotlyChart"] {
            overflow-x: auto !important;
            max-width: 100vw !important;
        }

        /* 甘特圖 Y 軸標籤優化 */
        .ytick text {
            max-width: 80px !important;
            overflow: hidden !important;
            text-overflow: ellipsis !important;
            white-space: nowrap !important;
        }

        /* Plotly Y 軸文字優化 */
        g.ytick text {
            font-size: 8px !important;
        }

        /* 按鈕和輸入框優化 */
        .stButton > button {
            width: 100%;
            font-size: 0.9rem;
        }

        .stTextInput > div > div > input {
            font-size: 0.9rem;
        }

        /* 卡片優化 */
        .metric-card {
            padding: 0.5rem;
            font-size: 0.9rem;
        }

        /* 側邊欄優化 */
        section[data-testid="stSidebar"] {
            width: 100% !important;
        }

        /* 文字大小調整 */
        h1 { font-size: 1.5rem !important; }
        h2 { font-size: 1.3rem !important; }
        h3 { font-size: 1.1rem !important; }
        h4 { font-size: 1rem !important; }

        /* 報告區塊優化 */
        .report-section {
            padding: 10px;
            font-size: 0.9rem;
        }

        /* 展開器優化 */
        div[data-testid="stExpander"] details summary p {
            font-size: 0.95rem;
        }
    }

    /* 小手機端 (< 480px) */
    @media only screen and (max-width: 480px) {
        .main-header {
            font-size: 1.2rem;
        }

        h1 { font-size: 1.2rem !important; }
        h2 { font-size: 1.1rem !important; }
        h3 { font-size: 1rem !important; }
        h4 { font-size: 0.95rem !important; }

        .stButton > button {
            font-size: 0.85rem;
            padding: 0.5rem;
        }
    }
</style>
""", unsafe_allow_html=True)


# ============================================================
# 資料載入與解析
# ============================================================
@st.cache_data
def load_excel_data(uploaded_file):
    """載入 Excel 檔案並解析各工作表"""
    try:
        from openpyxl import load_workbook

        xl = pd.ExcelFile(uploaded_file)
        sheet_names = xl.sheet_names

        # 使用 openpyxl 讀取格式資訊（背景色）
        uploaded_file.seek(0)  # 重置檔案指標
        wb = load_workbook(uploaded_file, data_only=False)
        ws_software = wb['軟體時程']

        # 讀取軟體時程表
        df_software = pd.read_excel(uploaded_file, sheet_name='軟體時程', header=None)
        
        # 提取專案資訊
        project_info = {
            'project_code': str(df_software.iloc[2, 2]) if pd.notna(df_software.iloc[2, 2]) else '',
            'project_name': str(df_software.iloc[3, 2]) if pd.notna(df_software.iloc[3, 2]) else '',
            'project_lead': str(df_software.iloc[4, 2]) if pd.notna(df_software.iloc[4, 2]) else '',
            'start_date': df_software.iloc[3, 9] if pd.notna(df_software.iloc[3, 9]) else None,
            'update_date': df_software.iloc[4, 12] if pd.notna(df_software.iloc[4, 12]) else None,
        }
        
        # 安全的數字轉換函數（定義在外層，避免重複定義）
        def safe_float(val, default=0):
            try:
                if pd.isna(val):
                    return default
                # 如果是字串且包含非數字字符（如標題），返回預設值
                if isinstance(val, str):
                    # 移除空白和換行
                    val_clean = str(val).strip()
                    # 檢查是否包含中文或其他非數字字符
                    if any(ord(c) > 127 for c in val_clean) or not val_clean.replace('.', '', 1).replace('-', '', 1).replace('+', '', 1).isdigit():
                        return default
                return float(val)
            except (ValueError, TypeError):
                return default

        def safe_int(val, default=0):
            try:
                if pd.isna(val):
                    return default
                if isinstance(val, str):
                    val_clean = str(val).strip()
                    if any(ord(c) > 127 for c in val_clean) or not val_clean.replace('-', '', 1).replace('+', '', 1).isdigit():
                        return default
                return int(float(val))
            except (ValueError, TypeError):
                return default

        def safe_datetime(val):
            """安全地轉換為日期時間，處理各種 Excel 日期格式（包含 2026/04/01(週三) 格式）"""
            try:
                if pd.isna(val):
                    return None

                # 如果是字串，嘗試移除括號中的中文（如：2026/04/01(週三) → 2026/04/01）
                if isinstance(val, str):
                    val_clean = str(val).strip()

                    # 移除括號及其內容（處理 "2026/04/01(週三)" 格式）
                    import re
                    val_clean = re.sub(r'\([^)]*\)', '', val_clean).strip()

                    # 如果清理後是空字串或只包含中文標題字樣，返回 None
                    if not val_clean or val_clean in ['計劃開始日期', '計劃完成日期', '實際開始日期', '實際完成日期']:
                        return None

                    # 使用清理後的字串進行轉換
                    val = val_clean

                # 嘗試轉換為 datetime
                result = pd.to_datetime(val, errors='coerce')

                # 如果轉換失敗，返回 None
                if pd.isna(result):
                    return None

                return result
            except:
                return None

        # 解析任務資料
        tasks = []
        filtered_count = 0  # 記錄被過濾的任務數量
        for i in range(6, len(df_software)):
            row = df_software.iloc[i]
            task_name = row[0]

            if pd.notna(task_name) and str(task_name).strip():
                # 跳過標題行（檢查是否 row[4] 包含 "百分比" 等關鍵字）
                if isinstance(row[4], str) and ('百分比' in str(row[4]) or '完成' in str(row[4])):
                    continue

                # 檢查 T 欄（備註欄，索引 19）是否包含「不支援」
                notes = str(row[19]).strip() if pd.notna(row[19]) else ''
                if '不支援' in notes:
                    filtered_count += 1
                    continue  # 跳過此任務，不顯示

                # ========== 判斷任務層級（多重方法）==========
                owner = str(row[2]) if pd.notna(row[2]) else ''
                task_name_str = str(task_name).strip()

                # 方法 1：檢查任務名稱是否有前導空格（Excel 中子項目可能縮排）
                has_leading_space = str(row[0]).startswith(' ') or str(row[0]).startswith('\t')

                # 方法 2：檢查 B 欄（索引 1）的層級標記
                level_marker = str(row[1]).strip() if pd.notna(row[1]) else ''

                # 判斷層級（支援多層級）
                level = 0  # 0=主項目, 1=次項目, 2=次次項目...
                if level_marker in ['主項目', '1', '大項', '大項目', 'parent', 'Parent']:
                    level = 0
                    is_parent_by_marker = True
                elif level_marker in ['次項目', '2', '子項', '子項目', 'child', 'Child']:
                    level = 1
                    is_parent_by_marker = False
                elif level_marker in ['次次項目', '3', '孫項', '孫項目']:
                    level = 2
                    is_parent_by_marker = False
                else:
                    is_parent_by_marker = False
                    # 如果沒有明確標記，嘗試從數字推斷層級
                    try:
                        level_num = int(level_marker)
                        if level_num > 0:
                            level = level_num - 1  # 1->0, 2->1, 3->2
                            is_parent_by_marker = (level_num == 1)
                    except:
                        pass

                # 方法 3：使用 Excel 背景色
                is_parent_by_color = False
                try:
                    excel_row = i + 1  # openpyxl 行索引從 1 開始
                    cell_a = ws_software.cell(row=excel_row, column=1)
                    if cell_a.fill and cell_a.fill.start_color:
                        color = cell_a.fill.start_color.rgb
                        if color and len(str(color)) >= 6:
                            color_str = str(color)[-6:]
                            try:
                                r = int(color_str[0:2], 16)
                                g = int(color_str[2:4], 16)
                                b = int(color_str[4:6], 16)
                                # 綠色：G > R 且 G > B，且 G > 150
                                is_parent_by_color = g > r and g > b and g > 150
                            except:
                                pass
                except:
                    pass

                # 方法 4：無負責單位 + 無日期
                has_dates = pd.notna(row[8]) and pd.notna(row[9])
                is_parent_by_logic = (not owner or owner.strip() == '') and not has_dates

                # 綜合判斷（優先級：層級標記 > 背景色 > 縮排 > 邏輯判斷）
                if is_parent_by_marker:
                    is_parent = True
                elif is_parent_by_color:
                    is_parent = True
                elif has_leading_space:
                    is_parent = False  # 有縮排 = 子項目
                else:
                    is_parent = is_parent_by_logic

                task = {
                    'id': len(tasks) + 1,
                    'row_index': i,
                    'task': str(task_name).strip(),
                    'is_parent': is_parent,  # 標記是否為大項目（主項目）
                    'level': level,  # 層級：0=主項目, 1=次項目, 2=次次項目
                    'owner': owner,
                    'progress_pct': safe_float(row[4]),
                    'target_pct': safe_float(row[5]),
                    'remaining_days': safe_int(row[6]),
                    'status': str(row[7]) if pd.notna(row[7]) else '',
                    'plan_start': safe_datetime(row[8]),
                    'plan_end': safe_datetime(row[9]),
                    'plan_days': safe_int(row[10]),
                    'actual_start': safe_datetime(row[11]),
                    'actual_end': safe_datetime(row[12]),
                    'actual_days': safe_int(row[13]),
                    'variance_days': safe_int(row[14]),
                    'coord_time': str(row[15]) if pd.notna(row[15]) else '',
                    'coord_manpower': str(row[16]) if pd.notna(row[16]) else '',
                    'coord_area': str(row[17]) if pd.notna(row[17]) else '',
                    'coord_equipment': str(row[18]) if pd.notna(row[18]) else '',
                    'notes': str(row[19]) if pd.notna(row[19]) else '',
                }
                tasks.append(task)
        
        df_tasks = pd.DataFrame(tasks)

        # 確保 progress_pct 為 0-100 格式
        if not df_tasks.empty and 'progress_pct' in df_tasks.columns:
            max_progress = df_tasks['progress_pct'].max()
            if max_progress <= 1 and max_progress > 0:
                df_tasks['progress_pct'] = df_tasks['progress_pct'] * 100
            if 'target_pct' in df_tasks.columns:
                max_target = df_tasks['target_pct'].max()
                if max_target <= 1 and max_target > 0:
                    df_tasks['target_pct'] = df_tasks['target_pct'] * 100

        # 自動計算狀態（如果 status 欄位為空）
        if not df_tasks.empty:
            today = datetime.now().date()
            def calc_status(row):
                if row['status'] and str(row['status']).strip():
                    return row['status']
                progress = row['progress_pct'] if pd.notna(row['progress_pct']) else 0
                plan_end = row['plan_end']
                if progress >= 100:
                    return 'Done'
                elif plan_end is not None and pd.notna(plan_end):
                    try:
                        end_date = plan_end.date() if hasattr(plan_end, 'date') else plan_end
                        if end_date < today:
                            return 'Delay'
                    except:
                        pass
                return 'Going'
            df_tasks['status'] = df_tasks.apply(calc_status, axis=1)

        # 讀取系統時程（支援 系統時程_C, 系統時程_A, 系統時程 等名稱）
        system_sheet_name = None
        for sn in sheet_names:
            if '系統時程' in sn:
                system_sheet_name = sn
                break

        if system_sheet_name:
            df_system = pd.read_excel(uploaded_file, sheet_name=system_sheet_name, header=None)
        else:
            df_system = pd.DataFrame()
        system_items = []
        current_area = ''
        current_main = ''

        # 找到階層欄位的索引（通常在第3欄或標題行含「階層」）
        hierarchy_col = 3  # 預設第4欄（索引3）
        if len(df_system) > 0:
            # 嘗試從標題行找到階層欄位
            for idx, val in enumerate(df_system.iloc[0]):
                if pd.notna(val) and '階層' in str(val):
                    hierarchy_col = idx
                    break

        for i in range(5, len(df_system)):
            row = df_system.iloc[i]
            item_name = str(row[0]).strip() if pd.notna(row[0]) else ''

            if item_name:
                # 讀取完成百分比（自動判斷 0-1 或 0-100 格式）
                pct = safe_float(row[2])
                if pct is not None and pct <= 1:
                    pct = pct * 100

                # 從階層欄位讀取項目類型
                hierarchy_value = str(row[hierarchy_col]).strip() if pd.notna(row[hierarchy_col]) else ''

                # 判斷項目類型: area, main, sub
                if '區域' in hierarchy_value or '區域' in item_name:
                    item_type = 'area'
                    current_area = item_name
                    current_main = ''
                elif '主項目' in hierarchy_value:
                    item_type = 'main'
                    current_main = item_name
                elif '次項目' in hierarchy_value:
                    item_type = 'sub'
                else:
                    # 備用邏輯：如果階層欄位為空，根據名稱判斷
                    if '區域' in item_name:
                        item_type = 'area'
                        current_area = item_name
                        current_main = ''
                    else:
                        item_type = 'sub'

                item = {
                    'area': current_area,
                    'main_item': current_main if item_type == 'sub' else ('' if item_type == 'area' else item_name),
                    'item': item_name.strip(),
                    'item_type': item_type,
                    'hierarchy': hierarchy_value,
                    'target_date': safe_datetime(row[1]),
                    'completion_pct': pct,
                    'is_area': item_type == 'area',
                    'is_main': item_type == 'main',
                }
                system_items.append(item)
        df_system_tasks = pd.DataFrame(system_items)
        
        # 讀取進度統計（包含「工作進度」的工作表）
        df_engineering = pd.DataFrame()
        progress_stats = []
        try:
            # 嘗試找到包含「工作進度」的工作表
            eng_sheet_name = None
            for sn in sheet_names:
                if '工作進度' in sn:
                    eng_sheet_name = sn
                    break

            if eng_sheet_name:
                df_eng_raw = pd.read_excel(uploaded_file, sheet_name=eng_sheet_name, header=None)
                df_engineering = df_eng_raw

                # 解析進度統計欄位
                # 欄位結構: 區域, 項目, C鋼(目標,實際), 軌道(目標,實際), HID(目標,實際),
                #          踩點圖資(目標,實際), Area Sensor(目標,實際), 走行提速(目標,實際),
                #          OHB(安裝,實際,教點,實際,Cycle,實際), Cycle Test(目標,實際),
                #          EQ Teaching(PIO安裝,教點), Hot Run, RTD Test, Release
                for i in range(2, len(df_eng_raw)):  # 跳過標題列
                    row = df_eng_raw.iloc[i]
                    area = str(row[0]).strip() if pd.notna(row[0]) else ''
                    item = str(row[1]).strip() if pd.notna(row[1]) else ''

                    if area or item:
                        stat = {
                            '區域': area,
                            '項目': item,
                            'C鋼_目標': safe_datetime(row[2]) if len(row) > 2 else None,
                            'C鋼_實際': safe_datetime(row[3]) if len(row) > 3 else None,
                            '軌道_目標': safe_datetime(row[4]) if len(row) > 4 else None,
                            '軌道_實際': safe_datetime(row[5]) if len(row) > 5 else None,
                            'HID_目標': safe_datetime(row[6]) if len(row) > 6 else None,
                            'HID_實際': safe_datetime(row[7]) if len(row) > 7 else None,
                            '踩點圖資_目標': safe_datetime(row[8]) if len(row) > 8 else None,
                            '踩點圖資_實際': safe_datetime(row[9]) if len(row) > 9 else None,
                            'AreaSensor_目標': safe_datetime(row[10]) if len(row) > 10 else None,
                            'AreaSensor_實際': safe_datetime(row[11]) if len(row) > 11 else None,
                            '走行提速_目標': safe_datetime(row[12]) if len(row) > 12 else None,
                            '走行提速_實際': safe_datetime(row[13]) if len(row) > 13 else None,
                            'OHB安裝_目標': safe_datetime(row[14]) if len(row) > 14 else None,
                            'OHB安裝_實際': safe_datetime(row[15]) if len(row) > 15 else None,
                            'OHB教點_目標': safe_datetime(row[16]) if len(row) > 16 else None,
                            'OHB教點_實際': safe_datetime(row[17]) if len(row) > 17 else None,
                            'OHBCycle_目標': safe_datetime(row[18]) if len(row) > 18 else None,
                            'OHBCycle_實際': safe_datetime(row[19]) if len(row) > 19 else None,
                            'CycleTest_目標': safe_datetime(row[20]) if len(row) > 20 else None,
                            'CycleTest_實際': safe_datetime(row[21]) if len(row) > 21 else None,
                            'EQTeaching_PIO安裝': safe_datetime(row[22]) if len(row) > 22 else None,
                            'EQTeaching_教點': safe_datetime(row[23]) if len(row) > 23 else None,
                            'HotRun': safe_datetime(row[24]) if len(row) > 24 else None,
                            'RTDTest': safe_datetime(row[25]) if len(row) > 25 else None,
                            'Release': safe_datetime(row[26]) if len(row) > 26 else None,
                        }
                        progress_stats.append(stat)
        except Exception as e:
            pass

        df_progress_stats = pd.DataFrame(progress_stats) if progress_stats else pd.DataFrame()
        
        # 讀取 EQ 工作清單
        try:
            df_eq = pd.read_excel(uploaded_file, sheet_name='EQ 工作清單', header=None)
        except:
            df_eq = pd.DataFrame()

        # 讀取 Layout 分頁的圖片
        layout_images = []
        try:
            if 'Layout' in sheet_names:
                ws_layout = wb['Layout']
                if hasattr(ws_layout, '_images') and ws_layout._images:
                    for img in ws_layout._images:
                        try:
                            # 提取圖片資訊
                            import io
                            # 獲取圖片二進制資料
                            if hasattr(img, 'ref') and hasattr(img.ref, 'getvalue'):
                                img_bytes = img.ref.getvalue()
                            elif hasattr(img, '_data'):
                                img_bytes = img._data()
                            else:
                                continue

                            if img_bytes:
                                layout_images.append(img_bytes)
                        except:
                            # 靜默跳過無法讀取的圖片
                            continue
        except:
            pass  # 如果沒有 Layout 分頁或無法讀取，就忽略

        return {
            'project_info': project_info,
            'tasks': df_tasks,
            'system_tasks': df_system_tasks,
            'engineering': df_engineering,
            'progress_stats': df_progress_stats,  # 進度統計
            'eq_list': df_eq,
            'raw_software': df_software,
            'sheet_names': sheet_names,
            'layout_images': layout_images,
            'filtered_count': filtered_count,  # 被過濾的任務數量
        }
    except Exception as e:
        st.error(f"載入檔案錯誤: {str(e)}")
        return None


# ============================================================
# 圖表生成函數
# ============================================================
def create_gantt_chart(df_tasks, show_actual=False, show_today_line=True, gantt_auto_range=True, enable_zoom=False):
    """建立甘特圖（使用 plotly.express.timeline）

    Args:
        df_tasks: 任務資料框
        show_actual: 是否顯示實際進度
        show_today_line: 是否顯示今日線
        gantt_auto_range: 是否自動範圍
        enable_zoom: 是否啟用縮放和拖曳（建議手機端開啟，電腦端關閉）
    """
    gantt_data = df_tasks[df_tasks['plan_start'].notna() & df_tasks['plan_end'].notna()].copy()

    if gantt_data.empty:
        return None

    color_map = {
        'Done': '#28a745',
        'Going': '#ffc107',
        'Delay': '#dc3545',
        '': '#6c757d'
    }

    try:
        # 準備資料給 px.timeline
        gantt_data['Start'] = pd.to_datetime(gantt_data['plan_start'])
        gantt_data['Finish'] = pd.to_datetime(gantt_data['plan_end'])
        # 保留完整任務名稱用於 hover
        gantt_data['TaskFull'] = gantt_data['task']
        # 縮短任務名稱以適應屏幕（手機端更短）
        # 如果啟用縮放（通常是手機端），使用更短的名稱
        max_chars = 10 if enable_zoom else 20
        gantt_data['Task'] = gantt_data['task'].apply(
            lambda x: str(x)[:max_chars] + '...' if len(str(x)) > max_chars else str(x)
        )
        gantt_data['Status'] = gantt_data['status']

        # 創建甘特圖
        import plotly.express as px
        fig = px.timeline(
            gantt_data,
            x_start='Start',
            x_end='Finish',
            y='Task',
            color='Status',
            color_discrete_map=color_map,
            title='📅 專案甘特圖',
            hover_data={'TaskFull': True, 'owner': True, 'Task': False},  # 在 hover 時顯示完整任務名稱
            labels={'TaskFull': '任務名稱'}
        )

        # 反轉 Y 軸，使第一個任務在最上面
        fig.update_yaxes(autorange='reversed')

        # 計算專案時間範圍
        min_date = gantt_data['Start'].min()
        max_date = gantt_data['Finish'].max()
        today = pd.Timestamp.now()

        # 根據設定決定 X 軸範圍
        if gantt_auto_range:
            # 自動範圍：只顯示專案時間範圍 + 5% 緩衝
            date_range = (max_date - min_date).total_seconds()
            buffer = pd.Timedelta(seconds=date_range * 0.05)
            x_range_start = min_date - buffer
            x_range_end = max_date + buffer
        else:
            # 完整範圍：從今日（或專案開始，取較早者）到專案結束
            x_range_start = min(today, min_date) - pd.Timedelta(days=7)
            x_range_end = max_date + pd.Timedelta(days=7)

        # 根據是否啟用縮放（通常代表手機端）調整邊距
        left_margin = 70 if enable_zoom else 120  # 手機端大幅減少左側邊距
        y_tickfont_size = 8 if enable_zoom else 9  # 手機端字體更小

        # 設定高度和 X 軸範圍
        fig.update_layout(
            height=max(500, len(gantt_data) * 28),
            xaxis_title='日期',
            yaxis_title='',
            xaxis_range=[x_range_start, x_range_end],
            # 優化顯示（手機端更緊湊）
            margin=dict(l=left_margin, r=20, t=50, b=50),
            font=dict(size=10),  # 縮小字體
            yaxis=dict(
                tickfont=dict(size=y_tickfont_size),  # Y軸標籤字體
                automargin=False,  # 關閉自動邊距，使用固定值
                tickmode='linear',  # 線性刻度
                side='left'  # 標籤在左側
            ),
            xaxis=dict(
                tickfont=dict(size=9),  # X軸標籤字體更小
                automargin=True
            ),
            # 根據設定啟用/禁用拖曳
            dragmode='pan' if enable_zoom else False,
            # 圖表標題字體
            title=dict(
                font=dict(size=14),
                x=0.5,  # 居中
                xanchor='center'
            )
        )

        # 配置交互選項（根據設定啟用/禁用縮放）
        fig.update_xaxes(fixedrange=not enable_zoom)  # enable_zoom=True 時允許縮放
        fig.update_yaxes(fixedrange=not enable_zoom)  # enable_zoom=True 時允許縮放

        # 顯示今日線（依據用戶設定）
        if show_today_line:
            try:
                # 如果是自動範圍模式，只在今日落在範圍內時顯示
                # 如果是完整範圍模式，總是顯示
                should_show = not gantt_auto_range or (x_range_start <= today <= x_range_end)

                if should_show:
                    fig.add_shape(
                        type="line",
                        x0=today, x1=today,
                        y0=0, y1=1,
                        yref="paper",
                        line=dict(color="red", width=2, dash="dash"),
                    )
                    fig.add_annotation(
                        x=today, y=1,
                        yref="paper",
                        text="今日",
                        showarrow=False,
                        yshift=10,
                        font=dict(color="red", size=12)
                    )
            except:
                pass

        return fig

    except Exception as e:
        # 如果使用 px.timeline 失敗，記錄錯誤
        if 'gantt_chart_error_info' not in st.session_state:
            st.session_state['gantt_chart_error_info'] = {
                'total': len(gantt_data),
                'success': 0,
                'error': len(gantt_data),
                'messages': [f'px.timeline 錯誤: {str(e)}']
            }
        return None

    # 實際時程（如果有）
    if show_actual:
        actual_data = gantt_data[gantt_data['actual_start'].notna() & gantt_data['actual_end'].notna()]
        for idx, row in actual_data.iterrows():
            try:
                actual_start = pd.to_datetime(row['actual_start'])
                actual_end = pd.to_datetime(row['actual_end'])

                fig.add_trace(go.Bar(
                    name='實際',
                    y=[row['task']],
                    x=[(actual_end - actual_start).days],
                    base=actual_start,
                    orientation='h',
                    marker_color='rgba(0,0,0,0.3)',
                    marker_line_color='black',
                    marker_line_width=2,
                    opacity=0.5,
                    showlegend=False,
                ))
            except Exception as e:
                continue

    # 設定版面配置
    fig.update_layout(
        title='📅 專案甘特圖',
        height=max(500, len(gantt_data) * 28),
        xaxis_title='日期',
        yaxis_title='',
        barmode='overlay',
        yaxis={'categoryorder': 'trace'},  # 改為 'trace' 以保持 Excel 中的順序（從上到下）
        yaxis_autorange='reversed',  # 反轉 y 軸，使第一個任務在最上面
        xaxis={'type': 'date'},
    )

    # 加入今日線（使用 add_shape 而不是 add_vline，避免日期格式問題）
    try:
        today = pd.Timestamp.now()
        fig.add_shape(
            type="line",
            x0=today, x1=today,
            y0=0, y1=1,
            yref="paper",
            line=dict(color="red", width=2, dash="dash"),
        )
        fig.add_annotation(
            x=today, y=1,
            yref="paper",
            text="今日",
            showarrow=False,
            yshift=10,
            font=dict(color="red", size=12)
        )
    except Exception as e:
        pass  # 如果加今日線失敗，就不加

    return fig


def create_status_pie(df_tasks):
    """狀態圓餅圖"""
    if df_tasks.empty:
        return None

    status_counts = df_tasks['status'].value_counts()

    if status_counts.empty:
        return None

    colors = {'Done': '#28a745', 'Going': '#ffc107', 'Delay': '#dc3545', '': '#6c757d'}

    fig = go.Figure(data=[go.Pie(
        labels=status_counts.index,
        values=status_counts.values,
        hole=0.4,
        marker_colors=[colors.get(s, '#6c757d') for s in status_counts.index],
        textinfo='value+percent',
        textposition='inside',
    )])

    fig.update_layout(title='📊 任務狀態分佈', height=350)
    return fig


def create_owner_workload(df_tasks):
    """負責單位工作量"""
    if df_tasks.empty:
        return None

    owner_stats = df_tasks.groupby('owner').agg({
        'task': 'count',
        'status': lambda x: list(x)
    }).reset_index()

    owner_stats['done'] = owner_stats['status'].apply(lambda x: x.count('Done'))
    owner_stats['going'] = owner_stats['status'].apply(lambda x: x.count('Going'))
    owner_stats['delay'] = owner_stats['status'].apply(lambda x: x.count('Delay'))
    owner_stats = owner_stats[owner_stats['owner'] != ''].sort_values('task', ascending=True)

    if owner_stats.empty:
        return None

    fig = go.Figure()
    fig.add_trace(go.Bar(name='已完成', y=owner_stats['owner'], x=owner_stats['done'],
                        orientation='h', marker_color='#28a745'))
    fig.add_trace(go.Bar(name='進行中', y=owner_stats['owner'], x=owner_stats['going'],
                        orientation='h', marker_color='#ffc107'))
    fig.add_trace(go.Bar(name='延遲', y=owner_stats['owner'], x=owner_stats['delay'],
                        orientation='h', marker_color='#dc3545'))

    fig.update_layout(
        barmode='stack',
        title='👥 各負責單位工作量',
        height=max(300, len(owner_stats) * 30),
        xaxis_title='任務數量',
    )
    return fig


def create_progress_trend(df_tasks):
    """進度趨勢圖（模擬）"""
    if df_tasks.empty:
        return None

    # 根據計劃完成日期模擬進度
    dates = pd.date_range(start='2025-05-01', end='2025-09-30', freq='W')

    progress_data = []
    for date in dates:
        done = len(df_tasks[(df_tasks['plan_end'].notna()) & (df_tasks['plan_end'] <= date)])
        total = len(df_tasks)
        progress_data.append({
            'date': date,
            'completed': done,
            'completion_rate': done / total * 100 if total > 0 else 0
        })

    df_progress = pd.DataFrame(progress_data)

    fig = make_subplots(specs=[[{"secondary_y": True}]])

    fig.add_trace(
        go.Bar(x=df_progress['date'], y=df_progress['completed'],
               name='累計完成數', marker_color='#28a745', opacity=0.7),
        secondary_y=False,
    )

    fig.add_trace(
        go.Scatter(x=df_progress['date'], y=df_progress['completion_rate'],
                  name='完成率 %', line=dict(color='#1f77b4', width=3)),
        secondary_y=True,
    )

    fig.update_layout(title='📈 進度趨勢圖', height=400)
    fig.update_yaxes(title_text="完成數量", secondary_y=False)
    fig.update_yaxes(title_text="完成率 (%)", secondary_y=True, range=[0, 100])

    return fig


def create_risk_matrix(df_tasks):
    """風險評估矩陣"""
    delay_tasks = df_tasks[df_tasks['status'] == 'Delay'].copy()
    
    if delay_tasks.empty:
        return None
    
    # 計算風險等級（基於誤差天數）
    def calc_risk(variance):
        if pd.isna(variance) or variance == 0:
            return 'low'
        elif abs(variance) <= 7:
            return 'medium'
        else:
            return 'high'
    
    delay_tasks['risk_level'] = delay_tasks['variance_days'].apply(calc_risk)
    
    risk_colors = {'high': '#dc3545', 'medium': '#ffc107', 'low': '#28a745'}
    
    fig = go.Figure()
    
    for risk in ['high', 'medium', 'low']:
        risk_data = delay_tasks[delay_tasks['risk_level'] == risk]
        if not risk_data.empty:
            fig.add_trace(go.Scatter(
                x=risk_data['variance_days'].abs(),
                y=risk_data['plan_days'],
                mode='markers+text',
                name=f'{risk.upper()} 風險',
                marker=dict(size=15, color=risk_colors[risk]),
                text=risk_data['task'].str[:15],
                textposition='top center',
                hovertemplate='<b>%{text}</b><br>誤差: %{x} 天<br>計劃天數: %{y} 天<extra></extra>'
            ))
    
    fig.update_layout(
        title='⚠️ 風險評估矩陣',
        xaxis_title='誤差天數（絕對值）',
        yaxis_title='計劃天數',
        height=400,
    )
    
    return fig


def create_progress_distribution(df_tasks):
    """進度區間分布圖"""
    if df_tasks.empty:
        return None

    ranges = [
        (0, 0, '未開始 (0%)', '#9e9e9e'),
        (1, 25, '剛開始 (1-25%)', '#ea4335'),
        (26, 50, '進行中 (26-50%)', '#f9ab00'),
        (51, 75, '過半 (51-75%)', '#1a73e8'),
        (76, 99, '接近完成 (76-99%)', '#34a853'),
        (100, 100, '已完成 (100%)', '#1e7e34')
    ]

    data = []
    for min_val, max_val, label, color in ranges:
        count = len(df_tasks[(df_tasks['progress_pct'] >= min_val) & (df_tasks['progress_pct'] <= max_val)])
        data.append({'range': label, 'count': count, 'color': color})

    df_dist = pd.DataFrame(data)

    fig = go.Figure(data=[
        go.Bar(
            x=df_dist['range'],
            y=df_dist['count'],
            marker_color=df_dist['color'],
            text=df_dist['count'],
            textposition='auto',
        )
    ])

    fig.update_layout(
        title='📊 任務進度區間分布',
        xaxis_title='進度區間',
        yaxis_title='任務數量',
        height=350,
    )
    return fig


def create_owner_progress_chart(df_tasks):
    """負責人平均進度圖"""
    if df_tasks.empty:
        return None

    owner_stats = df_tasks.groupby('owner').agg({
        'progress_pct': 'mean',
        'task': 'count'
    }).reset_index()

    owner_stats = owner_stats[owner_stats['owner'] != ''].sort_values('progress_pct', ascending=True)

    if owner_stats.empty:
        return None

    colors = []
    for pct in owner_stats['progress_pct']:
        if pct >= 80:
            colors.append('#34a853')
        elif pct >= 50:
            colors.append('#1a73e8')
        elif pct >= 25:
            colors.append('#f9ab00')
        else:
            colors.append('#ea4335')

    fig = go.Figure(data=[
        go.Bar(
            y=owner_stats['owner'],
            x=owner_stats['progress_pct'],
            orientation='h',
            marker_color=colors,
            text=[f'{p:.1f}%' for p in owner_stats['progress_pct']],
            textposition='auto',
        )
    ])

    fig.update_layout(
        title='👥 各負責人平均進度',
        xaxis_title='平均進度 (%)',
        height=max(300, len(owner_stats) * 35),
        xaxis=dict(range=[0, 100]),
    )
    return fig


def create_area_progress(df_system):
    """區域進度圖"""
    area_data = df_system[df_system['is_area'] == True].copy()
    
    if area_data.empty:
        return None
    
    fig = go.Figure()
    
    fig.add_trace(go.Bar(
        x=area_data['item'],
        y=area_data['completion_pct'] * 100,
        marker_color=area_data['completion_pct'].apply(
            lambda x: '#28a745' if x >= 0.7 else '#ffc107' if x >= 0.3 else '#dc3545'
        ),
        text=area_data['completion_pct'].apply(lambda x: f'{x*100:.0f}%'),
        textposition='outside',
    ))
    
    fig.update_layout(
        title='🏭 各區域完成進度',
        yaxis_title='完成率 (%)',
        yaxis_range=[0, 110],
        height=350,
    )
    
    return fig


# ============================================================
# 報表生成函數
# ============================================================
def generate_weekly_report(data, report_date=None):
    """生成週報"""
    if report_date is None:
        report_date = datetime.now()
    
    df_tasks = data['tasks']
    project_info = data['project_info']
    
    # 本週範圍
    week_start = report_date - timedelta(days=report_date.weekday())
    week_end = week_start + timedelta(days=6)
    
    # 統計數據
    total = len(df_tasks)
    done = len(df_tasks[df_tasks['status'] == 'Done'])
    going = len(df_tasks[df_tasks['status'] == 'Going'])
    delay = len(df_tasks[df_tasks['status'] == 'Delay'])
    
    # 本週完成的任務
    completed_this_week = df_tasks[
        (df_tasks['actual_end'].notna()) & 
        (df_tasks['actual_end'] >= week_start) & 
        (df_tasks['actual_end'] <= week_end)
    ]
    
    # 下週預計完成
    next_week_end = week_end + timedelta(days=7)
    planned_next_week = df_tasks[
        (df_tasks['plan_end'].notna()) & 
        (df_tasks['plan_end'] > week_end) & 
        (df_tasks['plan_end'] <= next_week_end) &
        (df_tasks['status'] != 'Done')
    ]
    
    report = f"""
# 📋 專案週報

**專案名稱：** {project_info['project_name']}  
**專案工令：** {project_info['project_code']}  
**報告日期：** {report_date.strftime('%Y-%m-%d')}  
**報告週期：** {week_start.strftime('%Y-%m-%d')} ~ {week_end.strftime('%Y-%m-%d')}

---

## 📊 整體進度概況

| 指標 | 數值 | 佔比 |
|------|------|------|
| 總任務數 | {total} | 100% |
| 已完成 | {done} | {done/total*100:.1f}% |
| 進行中 | {going} | {going/total*100:.1f}% |
| 延遲中 | {delay} | {delay/total*100:.1f}% |

**整體完成率：{done/total*100:.1f}%**

---

## ✅ 本週完成項目 ({len(completed_this_week)} 項)

"""
    
    if completed_this_week.empty:
        report += "本週無完成項目\n"
    else:
        for _, task in completed_this_week.iterrows():
            report += f"- {task['task']} ({task['owner']})\n"
    
    report += f"""
---

## 📅 下週計劃 ({len(planned_next_week)} 項)

"""
    
    if planned_next_week.empty:
        report += "下週無預計完成項目\n"
    else:
        for _, task in planned_next_week.iterrows():
            end_date = task['plan_end'].strftime('%m/%d') if pd.notna(task['plan_end']) else 'N/A'
            report += f"- {task['task']} (預計 {end_date}, {task['owner']})\n"
    
    report += f"""
---

## ⚠️ 風險與問題 ({delay} 項延遲)

"""
    
    delay_tasks = df_tasks[df_tasks['status'] == 'Delay']
    if delay_tasks.empty:
        report += "目前無延遲項目 ✅\n"
    else:
        for _, task in delay_tasks.head(10).iterrows():
            report += f"- **{task['task']}** - {task['owner']}\n"
    
    report += """
---

## 📝 備註

（請在此補充其他說明）

---
*此報告由 OHTC 專案管理儀表板自動生成*
"""
    
    return report


def generate_status_summary(data):
    """生成狀態摘要"""
    df_tasks = data['tasks']
    
    summary = {
        'total': len(df_tasks),
        'done': len(df_tasks[df_tasks['status'] == 'Done']),
        'going': len(df_tasks[df_tasks['status'] == 'Going']),
        'delay': len(df_tasks[df_tasks['status'] == 'Delay']),
        'delay_tasks': df_tasks[df_tasks['status'] == 'Delay'][['task', 'owner', 'plan_end', 'variance_days']].to_dict('records'),
        'upcoming': df_tasks[
            (df_tasks['status'] == 'Going') & 
            (df_tasks['plan_end'].notna()) &
            (df_tasks['plan_end'] <= datetime.now() + timedelta(days=7))
        ][['task', 'owner', 'plan_end']].to_dict('records'),
    }
    
    return summary


# ============================================================
# Excel 匯出函數
# ============================================================
def export_updated_excel(data, original_file, updated_tasks):
    """匯出更新後的 Excel（完整保留格式、公式、樣式）"""
    output = io.BytesIO()
    original_file.seek(0)

    # 載入工作簿，保留公式
    try:
        wb = load_workbook(original_file, keep_links=False, data_only=False)
    except:
        wb = load_workbook(original_file, keep_links=False)

    ws = wb['軟體時程']

    # 移除外部連結（但保留內部公式）
    if hasattr(wb, 'defined_names'):
        names_to_remove = []
        for name in wb.defined_names:
            try:
                if wb.defined_names[name].attr_text and '[' in str(wb.defined_names[name].attr_text):
                    names_to_remove.append(name)
            except:
                continue
        for name in names_to_remove:
            try:
                del wb.defined_names[name]
            except:
                continue

    if hasattr(wb, '_external_links'):
        wb._external_links = []

    # 只移除外部引用的公式，保留內部公式
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if cell.value.startswith('=') and '[' in cell.value and ']' in cell.value:
                        try:
                            cell.value = None
                        except:
                            continue

    # 更新專案資訊（保留格式）
    project_info = data.get('project_info', {})
    ws.cell(row=3, column=3).value = project_info.get('project_code', '')
    ws.cell(row=4, column=3).value = project_info.get('project_name', '')
    ws.cell(row=5, column=3).value = project_info.get('project_lead', '')

    # 獲取範本行（第 7 行）的樣式，用於新增任務
    template_row_idx = 7
    template_row_styles = {}
    for col in range(1, 21):
        cell = ws.cell(row=template_row_idx, column=col)
        template_row_styles[col] = {
            'font': cell.font.copy() if cell.font else None,
            'fill': cell.fill.copy() if cell.fill else None,
            'border': cell.border.copy() if cell.border else None,
            'alignment': cell.alignment.copy() if cell.alignment else None,
            'number_format': cell.number_format,
        }

    # 計算原始任務數量（假設從第 7 行開始）
    original_task_count = len(data.get('tasks', pd.DataFrame()))
    new_task_count = len(updated_tasks)

    # 如果任務數量減少，刪除多餘的行
    if new_task_count < original_task_count:
        for row_idx in range(7 + new_task_count, 7 + original_task_count):
            # 清空該行的內容，但保留格式
            for col in range(1, 21):
                ws.cell(row=row_idx, column=col).value = None

    # 更新或新增任務（只更新數值欄位，保留公式欄位）
    for idx, task in updated_tasks.iterrows():
        row_num = idx + 7  # 從第 7 行開始

        # 如果是新增的任務（超過原始行數），複製範本樣式
        if idx >= original_task_count:
            for col in range(1, 21):
                cell = ws.cell(row=row_num, column=col)
                style = template_row_styles.get(col, {})
                if style.get('font'):
                    cell.font = style['font']
                if style.get('fill'):
                    cell.fill = style['fill']
                if style.get('border'):
                    cell.border = style['border']
                if style.get('alignment'):
                    cell.alignment = style['alignment']
                if style.get('number_format'):
                    cell.number_format = style['number_format']

        # 只更新非公式欄位（保留 Excel 中的公式）
        # 欄位 1: 任務名稱
        cell = ws.cell(row=row_num, column=1)
        if not (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
            cell.value = task.get('task', '')

        # 欄位 3: 負責單位
        cell = ws.cell(row=row_num, column=3)
        if not (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
            cell.value = task.get('owner', '')

        # 欄位 5-7: 進度數值（可能有公式，檢查後再更新）
        for col, key in [(5, 'progress_pct'), (6, 'target_pct'), (7, 'remaining_days')]:
            cell = ws.cell(row=row_num, column=col)
            if not (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.value = task.get(key, 0)

        # 欄位 8: 狀態
        ws.cell(row=row_num, column=8).value = task.get('status', '')

        # 欄位 9-10: 計劃日期
        if pd.notna(task.get('plan_start')):
            ws.cell(row=row_num, column=9).value = pd.to_datetime(task['plan_start'])
        if pd.notna(task.get('plan_end')):
            ws.cell(row=row_num, column=10).value = pd.to_datetime(task['plan_end'])

        # 欄位 11: 計劃天數（可能是公式）
        cell = ws.cell(row=row_num, column=11)
        if not (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
            cell.value = task.get('plan_days', 0)

        # 欄位 12-13: 實際日期
        if pd.notna(task.get('actual_start')):
            ws.cell(row=row_num, column=12).value = pd.to_datetime(task['actual_start'])
        if pd.notna(task.get('actual_end')):
            ws.cell(row=row_num, column=13).value = pd.to_datetime(task['actual_end'])

        # 欄位 14-15: 實際天數、誤差天數（可能是公式）
        for col, key in [(14, 'actual_days'), (15, 'variance_days')]:
            cell = ws.cell(row=row_num, column=col)
            if not (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.value = task.get(key, 0)

        # 欄位 16-20: 協調欄位和備註
        ws.cell(row=row_num, column=16).value = task.get('coord_time', '')
        ws.cell(row=row_num, column=17).value = task.get('coord_manpower', '')
        ws.cell(row=row_num, column=18).value = task.get('coord_area', '')
        ws.cell(row=row_num, column=19).value = task.get('coord_equipment', '')
        ws.cell(row=row_num, column=20).value = task.get('notes', '')

    # 更新日期
    ws.cell(row=5, column=13).value = datetime.now()

    # 儲存
    wb.save(output)
    output.seek(0)
    return output


def export_report_to_word_format(report_content):
    """將報表匯出為可複製格式"""
    return report_content


def generate_export_filename(original_filename, project_name):
    """
    生成匯出檔案名稱
    格式：專案名稱+安裝排程表+_日期+_v版號(原版號+1)

    Args:
        original_filename: 原始上傳的檔案名稱
        project_name: 專案名稱

    Returns:
        新的檔案名稱字串
    """
    import re
    from datetime import datetime

    # 從原始檔案名提取版號
    version = 1
    if original_filename:
        # 嘗試匹配 _v數字 或 _V數字 格式
        version_match = re.search(r'_[vV](\d+)', original_filename)
        if version_match:
            version = int(version_match.group(1)) + 1
        else:
            # 如果沒有版號，檢查檔案名中是否有日期後的其他數字
            # 例如：PTI_PH2_OHTC安裝排程表_20251201.xlsx -> v1
            version = 1

    # 清理專案名稱（移除可能的特殊字符）
    clean_project_name = re.sub(r'[\\/:*?"<>|]', '', project_name) if project_name else 'OHTC'

    # 生成日期字串
    date_str = datetime.now().strftime('%Y%m%d')

    # 組合檔案名稱：專案名稱+安裝排程表+_日期+_v版號
    new_filename = f"{clean_project_name}_安裝排程表_{date_str}_v{version}.xlsx"

    return new_filename


# ============================================================
# 主應用程式
# ============================================================
def main():
    st.markdown('<h1 class="main-header">🏭 OHTC 專案管理儀表板 v2.0</h1>', unsafe_allow_html=True)
    
    # 側邊欄
    with st.sidebar:
        st.header("📁 檔案管理")
        uploaded_file = st.file_uploader(
            "上傳專案排程表",
            type=['xlsx', 'xls'],
            help="支援 OHTC 安裝排程表格式"
        )
        
        if uploaded_file:
            st.success(f"✅ {uploaded_file.name}")
        
        st.divider()
        
        st.header("⚙️ 顯示設定")
        show_actual = st.checkbox("顯示實際進度", value=True)
        show_completed = st.checkbox("顯示已完成項目", value=True)
        show_today_line = st.checkbox("顯示今日線", value=True, help="在甘特圖上標示今日位置")
        gantt_auto_range = st.checkbox(
            "甘特圖自動範圍",
            value=True,
            help="只顯示專案時間範圍，避免大片空白。取消勾選可看到從今日到專案的完整時間軸。"
        )

        # 智能縮放控制
        st.markdown("**📐 縮放控制**")

        # 初始化縮放設定（默認關閉，避免電腦端誤觸）
        if 'gantt_zoom_initialized' not in st.session_state:
            st.session_state['gantt_zoom_initialized'] = True
            # 檢測用戶代理（簡單方式：默認關閉）
            st.session_state['default_zoom_enabled'] = False

        enable_gantt_zoom = st.checkbox(
            "🔍 啟用甘特圖縮放/拖曳",
            value=st.session_state.get('default_zoom_enabled', False),
            help="📱 **手機模式（建議開啟）**：\n• 可雙指縮放、拖曳查看\n• 任務名稱縮短至10字\n• 左側邊距更小，圖表更大\n\n💻 **電腦模式（建議關閉）**：\n• 避免滾輪誤觸\n• 任務名稱顯示20字\n• 固定視圖",
            key="enable_gantt_zoom"
        )

        # 顯示當前狀態
        if enable_gantt_zoom:
            st.caption("✅ 手機模式：任務名稱短、可縮放拖曳")
        else:
            st.caption("🔒 電腦模式：任務名稱長、視圖鎖定")

        # 快速切換按鈕（讓手機用戶更方便）
        col_toggle1, col_toggle2 = st.columns(2)
        with col_toggle1:
            if st.button("📱 手機模式", use_container_width=True, help="縮短任務名稱、啟用縮放（建議手機）"):
                st.session_state['default_zoom_enabled'] = True
                st.rerun()
        with col_toggle2:
            if st.button("💻 電腦模式", use_container_width=True, help="較長任務名稱、鎖定縮放（建議電腦）"):
                st.session_state['default_zoom_enabled'] = False
                st.rerun()

        # Excel 原始資料檢視
        with st.expander("🔍 Excel 原始資料檢視（除錯用）", expanded=False):
            try:
                df_raw = pd.read_excel(uploaded_file, sheet_name='軟體時程', header=None, nrows=10)
                st.write("**Excel 前 10 行原始資料：**")
                st.dataframe(df_raw, use_container_width=True)
                st.caption("請確認第 8 欄（I 欄，0-based 索引）和第 9 欄（J 欄）是否為計劃開始/完成日期")
            except Exception as e:
                st.error(f"無法讀取原始資料：{e}")

        # 層級識別診斷（需要在上傳檔案後才顯示）
        if uploaded_file:
            with st.expander("🔬 層級識別診斷（Debug）", expanded=False):
                st.info("""
                💡 **如何改善層級識別？** 請參考 `EXCEL_FORMAT_GUIDE.md` 文檔

                **推薦方法：**
                1. **B 欄標記**：在 B 欄填入 `主項目`、`次項目` 標記層級
                2. **空格縮排**：子項目名稱前加 4 個空格
                3. **綠色背景**：大項目設定綠色背景（目前方式）
                """)

                try:
                    # 嘗試載入資料以顯示診斷
                    temp_data = load_excel_data(uploaded_file)
                    if temp_data and 'tasks' in temp_data:
                        temp_df = temp_data['tasks']

                        st.write("**前 10 個任務的層級判斷：**")
                        debug_data = []
                        level_names = {0: '主項目', 1: '次項目', 2: '次次項目'}

                        for idx, row in temp_df.head(10).iterrows():
                            level = row.get('level', 0)
                            level_display = level_names.get(level, f'層級{level+1}')

                            debug_data.append({
                                'ID': row['id'],
                                '任務名稱': row['task'][:30] + '...' if len(row['task']) > 30 else row['task'],
                                '層級': level_display,
                                '視覺化': f"{'  ' * level}{'■' if level == 0 else '├─'} {row['task'][:20]}"[:35],
                                '負責單位': (row['owner'][:10] + '...') if len(str(row['owner'])) > 10 else row['owner'] if row['owner'] else '(無)',
                                '有日期': '✅' if pd.notna(row['plan_start']) and pd.notna(row['plan_end']) else '❌'
                            })
                        st.dataframe(pd.DataFrame(debug_data), use_container_width=True)

                        st.caption("⚠️ 如果判斷不正確，請修改 Excel 格式（參考上方說明）或聯繫開發者")
                except Exception as e:
                    st.error(f"診斷工具載入失敗：{e}")

        st.divider()

        # 新專案範本生成器
        if TEMPLATE_GENERATOR_AVAILABLE:
            st.header("➕ 新專案範本")

            with st.expander("生成新專案 Excel 範本", expanded=False):
                st.markdown("**專案資訊：**")
                new_proj_name = st.text_input("專案名稱", value="新專案", key="new_proj_name")
                new_proj_code = st.text_input("專案工令", value="", key="new_proj_code")
                new_proj_lead = st.text_input("專案負責人", value="", key="new_proj_lead")
                new_proj_start = st.date_input("開始日期", value=datetime.now(), key="new_proj_start")

                if st.button("🔧 生成範本 Excel", type="primary", use_container_width=True):
                    try:
                        project_info = {
                            'name': new_proj_name,
                            'project_code': new_proj_code,
                            'lead': new_proj_lead,
                            'start_date': new_proj_start,
                        }

                        generator = ScheduleTemplateGenerator()

                        # 生成到 BytesIO 而非檔案
                        generator.create_software_schedule(project_info)
                        generator.create_system_schedule()
                        generator.create_engineering_progress()
                        generator.create_eq_list()
                        generator.create_location_map()
                        generator.create_fab_map()

                        # 儲存到 BytesIO
                        excel_buffer = io.BytesIO()
                        generator.wb.save(excel_buffer)
                        excel_buffer.seek(0)

                        st.download_button(
                            label="⬇️ 下載新專案範本",
                            data=excel_buffer,
                            file_name=f"{new_proj_name}_排程表_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        st.success("✅ 新專案範本已生成！")
                        st.info("💡 下載後可直接填寫任務資訊並上傳到儀表板")
                    except Exception as e:
                        st.error(f"❌ 生成失敗: {str(e)}")
                        st.exception(e)

            st.divider()

        st.header("📅 報表設定")
        report_date = st.date_input("報表日期", datetime.now())
    
    if uploaded_file is None:
        # 歡迎頁面
        st.info("👆 請先上傳專案排程表 Excel 檔案")
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            ### 🆕 v2.0 新功能
            
            - 📈 **進度趨勢分析** - 視覺化專案進展
            - ⚠️ **風險評估矩陣** - 識別高風險任務
            - 📋 **自動週報生成** - 一鍵產生報表
            - 📊 **區域進度追蹤** - 系統時程視覺化
            - ✏️ **任務編輯功能** - 直接更新狀態
            - 💾 **完整 Excel 匯出** - 保持原格式
            """)
        
        with col2:
            st.markdown("""
            ### 📋 支援格式

            - ✅ 軟體時程表（甘特圖）
            - ✅ 系統時程表（區域進度）
            - ✅ 工程進度確認表
            - ✅ EQ 工作清單

            ### 🚀 快速開始

            1. 上傳 Excel 排程表
            2. 瀏覽各項分析圖表
            3. 追蹤延遲項目
            4. 生成並下載報表

            ### 💡 智能過濾

            - 📝 T 欄（備註）標記「不支援」
            - 🔍 該任務自動隱藏不顯示
            - 📊 查看診斷資訊了解過濾數量
            """)
        return
    
    # 載入資料
    data = load_excel_data(uploaded_file)
    if data is None:
        return

    # 初始化 session_state（如果還沒有）
    if 'edited_project_info' not in st.session_state:
        st.session_state['edited_project_info'] = data['project_info'].copy()
    if 'edited_all_tasks' not in st.session_state:
        st.session_state['edited_all_tasks'] = data['tasks'].copy()
    if 'edited_system_tasks' not in st.session_state:
        st.session_state['edited_system_tasks'] = data['system_tasks'].copy()

    # 使用編輯後的資料（如果有），否則使用原始資料
    project_info = st.session_state.get('edited_project_info', data['project_info'])
    df_tasks = st.session_state.get('edited_all_tasks', data['tasks'])
    df_system = st.session_state.get('edited_system_tasks', data['system_tasks'])
    
    # 專案資訊卡
    st.markdown("### 📌 專案資訊")
    cols = st.columns(5)
    with cols[0]:
        st.metric("📋 專案工令", project_info['project_code'])
    with cols[1]:
        st.metric("🏭 專案名稱", project_info['project_name'][:15] + "...")
    with cols[2]:
        st.metric("👤 專案負責", project_info['project_lead'])
    with cols[3]:
        if project_info['start_date']:
            st.metric("📅 開始日期", pd.to_datetime(project_info['start_date']).strftime('%Y-%m-%d'))
    with cols[4]:
        total = len(df_tasks)
        done = len(df_tasks[df_tasks['status'] == 'Done'])
        st.metric("📊 完成率", f"{done/total*100:.1f}%", f"{done}/{total}")
    
    st.divider()
    
    # 關鍵指標卡
    total = len(df_tasks)
    done = len(df_tasks[df_tasks['status'] == 'Done'])
    going = len(df_tasks[df_tasks['status'] == 'Going'])
    delay = len(df_tasks[df_tasks['status'] == 'Delay'])
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #28a745, #20c997); padding: 20px; border-radius: 10px; color: white; text-align: center;">
            <div style="font-size: 2.5rem; font-weight: bold;">{done}</div>
            <div>✅ 已完成</div>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #ffc107, #fd7e14); padding: 20px; border-radius: 10px; color: white; text-align: center;">
            <div style="font-size: 2.5rem; font-weight: bold;">{going}</div>
            <div>🔄 進行中</div>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #dc3545, #c82333); padding: 20px; border-radius: 10px; color: white; text-align: center;">
            <div style="font-size: 2.5rem; font-weight: bold;">{delay}</div>
            <div>⚠️ 延遲中</div>
        </div>
        """, unsafe_allow_html=True)
    with col4:
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #6c757d, #495057); padding: 20px; border-radius: 10px; color: white; text-align: center;">
            <div style="font-size: 2.5rem; font-weight: bold;">{total}</div>
            <div>📝 總任務數</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.divider()
    
    # 主要標籤頁
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "📅 甘特圖",
        "📊 統計分析",
        "⚠️ 風險追蹤",
        "🏭 區域進度",
        "📋 進度統計",
        "✏️ 專案編輯",
        "📝 週報生成",
        "⬇️ 匯出"
    ])
    
    # Tab 1: 甘特圖
    with tab1:
        st.subheader("📅 專案甘特圖")

        # 使用提示（根據縮放設定顯示不同訊息）
        if enable_gantt_zoom:
            st.info("✅ **手機模式已啟用：** 可用滾輪/雙指縮放、拖曳查看甘特圖。任務名稱已縮短以節省空間（最多10字），點擊任務條可查看完整資訊。")
        else:
            st.info("🔒 **電腦模式（縮放鎖定）：** 視圖固定，避免誤觸。任務名稱最多顯示20字。如需啟用手機模式，請至側邊欄「⚙️ 顯示設定」。")

        # 診斷資訊
        total_tasks = len(df_tasks)
        tasks_with_dates = len(df_tasks[df_tasks['plan_start'].notna() & df_tasks['plan_end'].notna()])
        filtered_count = data.get('filtered_count', 0)  # 獲取被過濾的任務數量

        with st.expander("📊 資料診斷資訊", expanded=False):
            st.write(f"**顯示任務數：** {total_tasks}")
            st.write(f"**有計劃日期的任務：** {tasks_with_dates}")
            st.write(f"**缺少日期的任務：** {total_tasks - tasks_with_dates}")

            # 顯示過濾統計
            if filtered_count > 0:
                st.info(f"💡 **已自動過濾：** {filtered_count} 個任務（備註欄包含「不支援」）")
                st.caption("💡 如果任務的 T 欄（備註欄）包含「不支援」，該任務將不會在系統中顯示。")

            if tasks_with_dates == 0:
                st.error("⚠️ 所有任務都缺少計劃日期！請檢查 Excel 中的 I 欄（計劃開始）和 J 欄（計劃完成）是否有填寫日期。")

            # 顯示前 5 筆任務的日期狀態
            st.write("**前 5 筆任務的日期狀態：**")
            debug_df = df_tasks[['task', 'plan_start', 'plan_end', 'status']].head(5)
            st.dataframe(debug_df)

        gantt_fig = create_gantt_chart(df_tasks, show_actual, show_today_line, gantt_auto_range, enable_gantt_zoom)
        if gantt_fig:
            # 根據縮放設定配置 Plotly
            plotly_config = {
                'displayModeBar': True,  # 顯示工具列
                'modeBarButtonsToRemove': ['lasso2d', 'select2d'],  # 移除不常用的工具
                'displaylogo': False,  # 隱藏 Plotly logo
                'responsive': True  # 響應式
            }

            # 只在啟用縮放時添加 scrollZoom
            if enable_gantt_zoom:
                plotly_config['scrollZoom'] = True  # 啟用滾輪縮放
            else:
                plotly_config['scrollZoom'] = False  # 禁用滾輪縮放
                plotly_config['doubleClick'] = False  # 禁用雙擊重置

            st.plotly_chart(
                gantt_fig,
                use_container_width=True,
                config=plotly_config
            )
        else:
            st.warning("⚠️ 資料不足，無法生成甘特圖")
            st.info("💡 甘特圖需要任務包含「計劃開始日期」和「計劃完成日期」。請檢查 Excel 的 I 欄和 J 欄是否有填寫日期。")

            # 顯示錯誤詳情
            if 'gantt_chart_error_info' in st.session_state:
                error_info = st.session_state['gantt_chart_error_info']
                st.error(f"""
                **甘特圖生成失敗詳情：**
                - 有日期的任務數：{error_info['total']}
                - 成功處理：{error_info['success']}
                - 處理失敗：{error_info['error']}
                """)
                if error_info['messages']:
                    st.write("**前 3 個錯誤範例：**")
                    for msg in error_info['messages']:
                        st.write(f"- {msg}")
                # 清除錯誤訊息
                del st.session_state['gantt_chart_error_info']
    
    # Tab 2: 統計分析
    with tab2:
        # 子分頁
        sub_tab1, sub_tab2, sub_tab3 = st.tabs(["📊 任務狀態分布", "📈 進度趨勢圖", "👤 負責人分析"])

        with sub_tab1:
            col1, col2 = st.columns(2)
            with col1:
                status_fig = create_status_pie(df_tasks)
                if status_fig:
                    st.plotly_chart(status_fig, use_container_width=True)
                else:
                    st.warning("資料不足，無法生成狀態圓餅圖")
            with col2:
                owner_fig = create_owner_workload(df_tasks)
                if owner_fig:
                    st.plotly_chart(owner_fig, use_container_width=True)
                else:
                    st.warning("資料不足，無法生成負責單位工作量圖")

            st.divider()
            dist_fig = create_progress_distribution(df_tasks)
            if dist_fig:
                st.plotly_chart(dist_fig, use_container_width=True)

        with sub_tab2:
            trend_fig = create_progress_trend(df_tasks)
            if trend_fig:
                st.plotly_chart(trend_fig, use_container_width=True)
            else:
                st.warning("資料不足，無法生成進度趨勢圖")

        with sub_tab3:
            owner_progress_fig = create_owner_progress_chart(df_tasks)
            if owner_progress_fig:
                st.plotly_chart(owner_progress_fig, use_container_width=True)
            else:
                st.warning("資料不足，無法生成負責人進度圖")

            st.divider()
            if not df_tasks.empty:
                st.markdown("### 📋 負責人任務統計")
                owner_summary = df_tasks.groupby('owner').agg({
                    'task': 'count',
                    'progress_pct': 'mean',
                    'status': lambda x: (x == 'Done').sum()
                }).reset_index()
                owner_summary.columns = ['負責單位', '任務數', '平均進度(%)', '已完成']
                owner_summary['完成率(%)'] = (owner_summary['已完成'] / owner_summary['任務數'] * 100).round(1)
                owner_summary['平均進度(%)'] = owner_summary['平均進度(%)'].round(1)
                owner_summary = owner_summary[owner_summary['負責單位'] != ''].sort_values('任務數', ascending=False)
                st.dataframe(owner_summary, use_container_width=True, hide_index=True)
    
    # Tab 3: 風險追蹤
    with tab3:
        st.subheader("⚠️ 風險評估與追蹤")
        
        delay_df = df_tasks[df_tasks['status'] == 'Delay']
        
        if delay_df.empty:
            st.success("🎉 太棒了！目前沒有延遲項目！")
        else:
            st.error(f"⚠️ 共有 {len(delay_df)} 個延遲項目需要關注")
            
            col1, col2 = st.columns([2, 1])
            
            with col1:
                risk_fig = create_risk_matrix(df_tasks)
                if risk_fig:
                    st.plotly_chart(risk_fig, use_container_width=True)
            
            with col2:
                st.markdown("### 🔴 高風險項目")
                high_risk = delay_df[delay_df['variance_days'].abs() > 7]
                for _, task in high_risk.iterrows():
                    with st.expander(f"🔴 {task['task'][:30]}..."):
                        st.write(f"**負責單位:** {task['owner']}")
                        st.write(f"**誤差天數:** {task['variance_days']} 天")
                        if pd.notna(task['plan_end']):
                            st.write(f"**計劃完成:** {task['plan_end'].strftime('%Y-%m-%d')}")
            
            st.divider()
            
            # 延遲項目清單
            st.markdown("### 📋 完整延遲項目清單")
            st.dataframe(
                delay_df[['task', 'owner', 'plan_end', 'variance_days', 'notes']].rename(columns={
                    'task': '任務', 'owner': '負責單位', 'plan_end': '計劃完成',
                    'variance_days': '誤差天數', 'notes': '備註'
                }),
                use_container_width=True,
                hide_index=True,
            )
    
    # Tab 4: 區域進度
    with tab4:
        st.subheader("🏭 系統時程 - 區域進度")

        area_fig = create_area_progress(df_system)
        if area_fig:
            st.plotly_chart(area_fig, use_container_width=True)

        st.divider()

        # 各區域詳細進度（主項目/次項目分開顯示）
        areas = df_system[df_system['is_area'] == True]['item'].unique()

        for area in areas:
            with st.expander(f"📍 {area}"):
                area_items = df_system[(df_system['area'] == area) & (df_system['is_area'] == False)]
                if not area_items.empty:
                    # 取得該區域的主項目
                    main_items = area_items[area_items['is_main'] == True]['item'].unique()

                    for main_item in main_items:
                        # 主項目標題
                        main_row = area_items[area_items['item'] == main_item].iloc[0]
                        main_pct = main_row['completion_pct'] if pd.notna(main_row['completion_pct']) else 0
                        main_color = '#28a745' if main_pct >= 70 else '#ffc107' if main_pct >= 30 else '#dc3545'

                        st.markdown(f"""
                        <div style="background: #f8f9fa; padding: 10px; border-radius: 8px; margin: 10px 0 5px 0; border-left: 4px solid {main_color};">
                            <div style="display: flex; align-items: center;">
                                <div style="font-weight: bold; font-size: 1.1em; width: 250px;">📌 {main_item[:40]}</div>
                                <div style="flex: 1; background: #e9ecef; border-radius: 4px; height: 22px; margin: 0 10px;">
                                    <div style="width: {main_pct}%; background: {main_color}; height: 100%; border-radius: 4px;"></div>
                                </div>
                                <div style="width: 60px; text-align: right; font-weight: bold;">{main_pct:.0f}%</div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

                        # 該主項目下的次項目
                        sub_items = area_items[(area_items['main_item'] == main_item) & (area_items['is_main'] == False)]
                        if not sub_items.empty:
                            for _, sub_row in sub_items.iterrows():
                                sub_pct = sub_row['completion_pct'] if pd.notna(sub_row['completion_pct']) else 0
                                sub_color = '#28a745' if sub_pct >= 70 else '#ffc107' if sub_pct >= 30 else '#dc3545'
                                st.markdown(f"""
                                <div style="display: flex; align-items: center; margin: 3px 0; padding-left: 30px;">
                                    <div style="width: 220px; color: #666;">└ {sub_row['item'][:35]}</div>
                                    <div style="flex: 1; background: #e9ecef; border-radius: 4px; height: 16px; margin: 0 10px;">
                                        <div style="width: {sub_pct}%; background: {sub_color}; height: 100%; border-radius: 4px;"></div>
                                    </div>
                                    <div style="width: 60px; text-align: right; color: #666;">{sub_pct:.0f}%</div>
                                </div>
                                """, unsafe_allow_html=True)

                    # 處理沒有主項目的次項目（直接屬於區域的項目）
                    orphan_items = area_items[(area_items['main_item'] == '') & (area_items['is_main'] == False)]
                    if not orphan_items.empty:
                        st.markdown("<div style='margin-top: 15px;'><strong>其他項目：</strong></div>", unsafe_allow_html=True)
                        for _, item in orphan_items.iterrows():
                            pct = item['completion_pct'] if pd.notna(item['completion_pct']) else 0
                            color = '#28a745' if pct >= 70 else '#ffc107' if pct >= 30 else '#dc3545'
                            st.markdown(f"""
                            <div style="display: flex; align-items: center; margin: 5px 0;">
                                <div style="width: 200px;">{item['item'][:30]}</div>
                                <div style="flex: 1; background: #e9ecef; border-radius: 4px; height: 20px; margin: 0 10px;">
                                    <div style="width: {pct}%; background: {color}; height: 100%; border-radius: 4px;"></div>
                                </div>
                                <div style="width: 50px; text-align: right;">{pct:.0f}%</div>
                            </div>
                            """, unsafe_allow_html=True)
    
    # Tab 5: 進度統計
    with tab5:
        st.subheader("📋 進度統計")

        df_progress = data.get('progress_stats', pd.DataFrame())

        if df_progress.empty:
            st.warning("⚠️ 未找到進度統計資料（需要包含「工作進度」的工作表）")
        else:
            items_row1 = ['C鋼', '軌道', 'HID', '踩點圖資', 'AreaSensor', '走行提速']
            items_row2 = ['OHB安裝', 'OHB教點', 'OHBCycle', 'CycleTest']
            all_items = items_row1 + items_row2

            # 檢查是否有區域欄位
            if '區域' in df_progress.columns:
                areas = [a for a in df_progress['區域'].unique() if a and str(a).strip()]

                # 按區域分開統計
                st.markdown("### 📊 各區域完成統計")

                for area in areas:
                    area_data = df_progress[df_progress['區域'] == area]
                    area_count = len(area_data)

                    with st.expander(f"📍 {area} ({area_count} 項)", expanded=True):
                        # 第一排統計
                        cols1 = st.columns(len(items_row1))
                        for idx, item in enumerate(items_row1):
                            target_col = f'{item}_目標'
                            actual_col = f'{item}_實際'
                            if target_col in df_progress.columns and actual_col in df_progress.columns:
                                total = area_data[target_col].notna().sum()
                                done = area_data[actual_col].notna().sum()
                                pct = (done / total * 100) if total > 0 else 0
                                delta_color = "normal" if pct >= 50 else "inverse"
                                with cols1[idx]:
                                    st.metric(item, f"{done}/{total}", f"{pct:.0f}%", delta_color=delta_color if pct < 50 else "off")

                        # 第二排統計
                        cols2 = st.columns(len(items_row2) + 1)
                        for idx, item in enumerate(items_row2):
                            target_col = f'{item}_目標'
                            actual_col = f'{item}_實際'
                            if target_col in df_progress.columns and actual_col in df_progress.columns:
                                total = area_data[target_col].notna().sum()
                                done = area_data[actual_col].notna().sum()
                                pct = (done / total * 100) if total > 0 else 0
                                with cols2[idx]:
                                    st.metric(item, f"{done}/{total}", f"{pct:.0f}%")

                        # EQ Teaching 特殊處理
                        with cols2[len(items_row2)]:
                            pio_col = 'EQTeaching_PIO安裝'
                            teach_col = 'EQTeaching_教點'
                            if pio_col in df_progress.columns:
                                pio_done = area_data[pio_col].notna().sum()
                                teach_done = area_data[teach_col].notna().sum() if teach_col in df_progress.columns else 0
                                st.metric("EQ Teaching", f"PIO:{pio_done} 教點:{teach_done}")

                        # 區域內進度條
                        st.markdown("**進度條：**")
                        for item in all_items:
                            target_col = f'{item}_目標'
                            actual_col = f'{item}_實際'
                            if target_col in df_progress.columns and actual_col in df_progress.columns:
                                total = area_data[target_col].notna().sum()
                                done = area_data[actual_col].notna().sum()
                                pct = (done / total * 100) if total > 0 else 0
                                color = '#28a745' if pct >= 70 else '#ffc107' if pct >= 30 else '#dc3545'
                                st.markdown(f"""
                                <div style="display: flex; align-items: center; margin: 5px 0;">
                                    <div style="width: 100px; font-size: 0.9em;">{item}</div>
                                    <div style="flex: 1; background: #e9ecef; border-radius: 4px; height: 18px; margin: 0 10px;">
                                        <div style="width: {pct}%; background: {color}; height: 100%; border-radius: 4px;"></div>
                                    </div>
                                    <div style="width: 70px; text-align: right; font-size: 0.9em;">{done}/{total} ({pct:.0f}%)</div>
                                </div>
                                """, unsafe_allow_html=True)

                        # 區域明細表格
                        st.markdown("**項目明細：**")
                        display_cols = ['項目'] + [f'{item}_實際' for item in all_items if f'{item}_實際' in df_progress.columns]
                        if display_cols:
                            st.dataframe(area_data[display_cols], use_container_width=True, hide_index=True)

                st.divider()

                # 全部區域總計
                st.markdown("### 📈 全區域總計")
                total_cols = st.columns(len(all_items))
                for idx, item in enumerate(all_items):
                    target_col = f'{item}_目標'
                    actual_col = f'{item}_實際'
                    if target_col in df_progress.columns and actual_col in df_progress.columns:
                        total = df_progress[target_col].notna().sum()
                        done = df_progress[actual_col].notna().sum()
                        pct = (done / total * 100) if total > 0 else 0
                        with total_cols[idx]:
                            st.metric(item, f"{done}/{total}", f"{pct:.0f}%")

            else:
                # 沒有區域欄位時的備用顯示
                st.markdown("### 📊 各項目完成統計")
                cols1 = st.columns(len(items_row1))
                for idx, item in enumerate(items_row1):
                    target_col = f'{item}_目標'
                    actual_col = f'{item}_實際'
                    if target_col in df_progress.columns and actual_col in df_progress.columns:
                        total = df_progress[target_col].notna().sum()
                        done = df_progress[actual_col].notna().sum()
                        pct = (done / total * 100) if total > 0 else 0
                        with cols1[idx]:
                            st.metric(item, f"{done}/{total}", f"{pct:.0f}%")

            st.divider()

            # 完整資料表格
            st.markdown("### 📋 完整資料表格")
            st.dataframe(df_progress, use_container_width=True, height=400)

    # Tab 6: 專案編輯
    with tab6:
        st.subheader("✏️ 專案與任務編輯器")

        # 提示：篩選與操作說明
        st.info("💡 **使用提示：** 篩選與搜尋本身不會觸發頁面刷新。但執行操作（如新增、批量修改、儲存變更）後會重新載入頁面，此時會回到甘特圖分頁（這是 Streamlit 的限制）。修改完成後請前往「匯出」分頁儲存變更。")

        # 初始化編輯歷史（用於撤銷/重做）
        if 'edit_history' not in st.session_state:
            st.session_state['edit_history'] = []
            st.session_state['history_index'] = -1

        # 顯示編輯狀態
        status_col1, status_col2, status_col3 = st.columns([2, 1, 1])
        with status_col1:
            if 'last_edit_time' in st.session_state:
                st.info(f"💡 最後編輯時間：{st.session_state['last_edit_time']}｜所有分頁已同步更新")
        with status_col2:
            # 撤銷按鈕
            can_undo = st.session_state['history_index'] > 0
            if st.button("↶ 撤銷", disabled=not can_undo, use_container_width=True, help="返回上一步操作"):
                if can_undo:
                    st.session_state['history_index'] -= 1
                    st.session_state['edited_all_tasks'] = st.session_state['edit_history'][st.session_state['history_index']].copy()
                    st.success("✅ 已撤銷上一步操作")
                    st.rerun()
        with status_col3:
            # 重做按鈕
            can_redo = st.session_state['history_index'] < len(st.session_state['edit_history']) - 1
            if st.button("↷ 重做", disabled=not can_redo, use_container_width=True, help="重做已撤銷的操作"):
                if can_redo:
                    st.session_state['history_index'] += 1
                    st.session_state['edited_all_tasks'] = st.session_state['edit_history'][st.session_state['history_index']].copy()
                    st.success("✅ 已重做操作")
                    st.rerun()

        # 專案資訊編輯
        st.markdown("### 📌 專案資訊")
        with st.expander("點擊編輯專案資訊", expanded=False):
            col1, col2 = st.columns(2)
            with col1:
                new_project_code = st.text_input("專案工令", value=st.session_state['edited_project_info'].get('project_code', ''))
                new_project_name = st.text_input("專案名稱", value=st.session_state['edited_project_info'].get('project_name', ''))
            with col2:
                new_project_lead = st.text_input("專案負責人", value=st.session_state['edited_project_info'].get('project_lead', ''))
                new_start_date = st.date_input("開始日期", value=pd.to_datetime(st.session_state['edited_project_info'].get('start_date')) if pd.notna(st.session_state['edited_project_info'].get('start_date')) else datetime.now())

            if st.button("💾 更新專案資訊", key="update_project"):
                st.session_state['edited_project_info']['project_code'] = new_project_code
                st.session_state['edited_project_info']['project_name'] = new_project_name
                st.session_state['edited_project_info']['project_lead'] = new_project_lead
                st.session_state['edited_project_info']['start_date'] = new_start_date
                st.session_state['last_edit_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                st.success("✅ 專案資訊已更新｜所有圖表已同步")
                st.rerun()

        st.divider()

        # 任務編輯
        st.markdown("### 📋 任務清單編輯")

        # 篩選器
        st.markdown("**🔍 篩選與搜尋：**")
        filter_col1, filter_col2, filter_col3, filter_col4, filter_col5 = st.columns(5)
        with filter_col1:
            status_filter_edit = st.multiselect(
                "篩選狀態",
                options=['Done', 'Going', 'Delay'],
                default=['Done', 'Going', 'Delay'],
                key="status_filter_edit"
            )
        with filter_col2:
            # 安全地獲取負責單位列表（移除 NaN 和空值）
            owners_list = sorted([str(x) for x in st.session_state['edited_all_tasks']['owner'].dropna().unique() if str(x).strip()])
            owner_filter_edit = st.multiselect("篩選負責單位", options=owners_list, key="owner_filter_edit")
        with filter_col3:
            # 主項目篩選
            parent_filter_edit = st.selectbox(
                "篩選主項目",
                options=['全部', '僅主項目', '僅次項目'],
                index=0,
                key="parent_filter_edit"
            )
        with filter_col4:
            search_edit = st.text_input("🔍 搜尋任務關鍵字", key="search_edit")
        with filter_col5:
            if st.button("🔄 清除篩選", use_container_width=True):
                # 清除篩選條件（透過設定 key 的方式強制重設）
                for key in ['status_filter_edit', 'owner_filter_edit', 'parent_filter_edit', 'search_edit']:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()

        st.divider()

        # ========== 先定義篩選條件和變數 ==========
        # 套用篩選條件
        filtered_tasks = st.session_state['edited_all_tasks'].copy()

        # 篩選狀態
        if status_filter_edit:
            filtered_tasks = filtered_tasks[filtered_tasks['status'].isin(status_filter_edit)]

        # 篩選負責單位
        if owner_filter_edit:
            filtered_tasks = filtered_tasks[filtered_tasks['owner'].isin(owner_filter_edit)]

        # 篩選主項目
        if parent_filter_edit == '僅主項目':
            filtered_tasks = filtered_tasks[filtered_tasks['is_parent'] == True]
        elif parent_filter_edit == '僅次項目':
            filtered_tasks = filtered_tasks[filtered_tasks['is_parent'] == False]

        # 搜尋任務關鍵字
        if search_edit:
            filtered_tasks = filtered_tasks[
                filtered_tasks['task'].str.contains(search_edit, case=False, na=False) |
                filtered_tasks['notes'].str.contains(search_edit, case=False, na=False)
            ]

        # 獲取所有現有的負責單位（用於下拉選單）
        existing_owners = [str(x) for x in st.session_state['edited_all_tasks']['owner'].dropna().unique() if str(x).strip()]
        # 加入常用單位作為預設選項
        common_owners = ['TIM SMA', 'TIM Controls', 'TIM Mechanical', 'TIM Electrical', 'Vendor']
        owner_options = sorted(list(set(existing_owners + common_owners)))

        # ========== 操作按鈕與批量操作 ==========
        st.markdown("**操作：**")
        op_col1, op_col2, op_col3, op_col4, op_col5 = st.columns(5)

        with op_col1:
            if st.button("➕ 新增任務", type="primary", use_container_width=True):
                new_task = {
                    'id': len(st.session_state['edited_all_tasks']) + 1,
                    'row_index': len(st.session_state['edited_all_tasks']) + 6,
                    'task': '新任務',
                    'is_parent': False,  # 預設為子項目
                    'level': 1,  # 預設為次項目
                    'owner': '',
                    'progress_pct': 0,
                    'target_pct': 0,
                    'remaining_days': 0,
                    'status': 'Going',
                    'plan_start': pd.Timestamp.now(),
                    'plan_end': pd.Timestamp.now() + pd.Timedelta(days=7),
                    'plan_days': 7,
                    'actual_start': None,
                    'actual_end': None,
                    'actual_days': 0,
                    'variance_days': 0,
                    'coord_time': '',
                    'coord_manpower': '',
                    'coord_area': '',
                    'coord_equipment': '',
                    'notes': '',
                }
                st.session_state['edited_all_tasks'] = pd.concat([
                    st.session_state['edited_all_tasks'],
                    pd.DataFrame([new_task])
                ], ignore_index=True)
                st.rerun()

        # 批量操作區域
        with op_col2:
            with st.popover("📝 批量修改狀態", use_container_width=True):
                batch_status = st.selectbox("選擇新狀態", ["Done", "Going", "Delay"], key="batch_status")
                batch_task_ids = st.multiselect(
                    "選擇要修改的任務 ID",
                    options=filtered_tasks['id'].tolist(),
                    key="batch_status_ids"
                )
                if st.button("✅ 套用批量狀態修改", use_container_width=True):
                    if batch_task_ids:
                        for task_id in batch_task_ids:
                            idx = st.session_state['edited_all_tasks'][st.session_state['edited_all_tasks']['id'] == task_id].index
                            if len(idx) > 0:
                                st.session_state['edited_all_tasks'].loc[idx[0], 'status'] = batch_status
                        st.session_state['last_edit_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        st.success(f"✅ 已將 {len(batch_task_ids)} 個任務狀態改為 {batch_status}")
                        st.rerun()
                    else:
                        st.warning("⚠️ 請選擇至少一個任務")

        with op_col3:
            with st.popover("👥 批量修改負責單位", use_container_width=True):
                batch_owner = st.selectbox("選擇新負責單位", owner_options, key="batch_owner")
                batch_owner_ids = st.multiselect(
                    "選擇要修改的任務 ID",
                    options=filtered_tasks['id'].tolist(),
                    key="batch_owner_ids"
                )
                if st.button("✅ 套用批量負責單位修改", use_container_width=True):
                    if batch_owner_ids:
                        for task_id in batch_owner_ids:
                            idx = st.session_state['edited_all_tasks'][st.session_state['edited_all_tasks']['id'] == task_id].index
                            if len(idx) > 0:
                                st.session_state['edited_all_tasks'].loc[idx[0], 'owner'] = batch_owner
                        st.session_state['last_edit_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        st.success(f"✅ 已將 {len(batch_owner_ids)} 個任務負責單位改為 {batch_owner}")
                        st.rerun()
                    else:
                        st.warning("⚠️ 請選擇至少一個任務")

        with op_col4:
            with st.popover("🗑️ 批量刪除", use_container_width=True):
                batch_delete_ids = st.multiselect(
                    "選擇要刪除的任務 ID",
                    options=filtered_tasks['id'].tolist(),
                    key="batch_delete_ids"
                )
                st.warning(f"⚠️ 將刪除 {len(batch_delete_ids)} 個任務，此操作無法復原")
                if st.button("🗑️ 確認批量刪除", type="secondary", use_container_width=True):
                    if batch_delete_ids:
                        st.session_state['edited_all_tasks'] = st.session_state['edited_all_tasks'][
                            ~st.session_state['edited_all_tasks']['id'].isin(batch_delete_ids)
                        ].reset_index(drop=True)
                        # 重新計算 ID
                        st.session_state['edited_all_tasks']['id'] = range(1, len(st.session_state['edited_all_tasks']) + 1)
                        st.session_state['last_edit_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        st.success(f"✅ 已刪除 {len(batch_delete_ids)} 個任務")
                        st.rerun()
                    else:
                        st.warning("⚠️ 請選擇至少一個任務")

        with op_col5:
            with st.popover("📋 複製任務", use_container_width=True):
                copy_task_id = st.selectbox(
                    "選擇要複製的任務 ID",
                    options=filtered_tasks['id'].tolist(),
                    key="copy_task_id"
                )
                copy_count = st.number_input("複製份數", min_value=1, max_value=10, value=1, key="copy_count")

                if st.button("📋 確認複製", use_container_width=True):
                    if copy_task_id:
                        # 找到要複製的任務
                        original_task = st.session_state['edited_all_tasks'][
                            st.session_state['edited_all_tasks']['id'] == copy_task_id
                        ]

                        if not original_task.empty:
                            # 複製任務
                            for i in range(copy_count):
                                new_task = original_task.iloc[0].to_dict()
                                new_task['id'] = len(st.session_state['edited_all_tasks']) + 1
                                new_task['row_index'] = len(st.session_state['edited_all_tasks']) + 6
                                new_task['task'] = f"{new_task['task']} (副本{i+1})"

                                st.session_state['edited_all_tasks'] = pd.concat([
                                    st.session_state['edited_all_tasks'],
                                    pd.DataFrame([new_task])
                                ], ignore_index=True)

                            # 重新計算 ID
                            st.session_state['edited_all_tasks']['id'] = range(1, len(st.session_state['edited_all_tasks']) + 1)

                            st.session_state['last_edit_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            st.success(f"✅ 已複製 {copy_count} 個任務")
                            st.rerun()
                        else:
                            st.error("❌ 找不到要複製的任務")

        # 顯示選項
        show_all = st.checkbox("顯示所有欄位", value=False)

        # 顯示篩選結果數量
        st.caption(f"📊 顯示 {len(filtered_tasks)} / {len(st.session_state['edited_all_tasks'])} 個任務")

        # 檢查是否有篩選結果
        if filtered_tasks.empty:
            st.warning("⚠️ 沒有符合篩選條件的任務")
            st.stop()

        # 準備顯示用的資料（加入層級標記，與 Excel 一致）
        display_tasks = filtered_tasks.copy()

        def format_task_with_level(row):
            """根據層級格式化任務名稱"""
            try:
                level = row.get('level', 0) if hasattr(row, 'get') else 0
                task_name = str(row['task']) if 'task' in row else ''

                if level == 0:
                    # 主項目
                    return f"■ {task_name}"
                elif level == 1:
                    # 次項目
                    return f"  ├─ {task_name}"
                elif level == 2:
                    # 次次項目
                    return f"    └─ {task_name}"
                else:
                    # 更深層級
                    indent = "  " * level
                    return f"{indent}└─ {task_name}"
            except Exception as e:
                # 如果格式化失敗，返回原始任務名稱
                return str(row.get('task', '')) if hasattr(row, 'get') else ''

        # 安全地應用格式化函數
        try:
            display_tasks['task_display'] = display_tasks.apply(format_task_with_level, axis=1)
        except Exception as e:
            # 如果應用失敗，使用原始任務名稱
            display_tasks['task_display'] = display_tasks['task'].astype(str)

        # 可編輯的任務表格
        if show_all:
            # 顯示所有欄位
            edit_columns = ['id', 'task_display', 'owner', 'status', 'plan_start', 'plan_end',
                          'plan_days', 'actual_start', 'actual_end', 'progress_pct',
                          'variance_days', 'notes']
            column_names = {
                'id': 'ID', 'task_display': '任務名稱', 'owner': '負責單位', 'status': '狀態',
                'plan_start': '計劃開始', 'plan_end': '計劃完成', 'plan_days': '計劃天數',
                'actual_start': '實際開始', 'actual_end': '實際完成',
                'progress_pct': '完成%', 'variance_days': '誤差天數', 'notes': '備註'
            }
        else:
            # 只顯示主要欄位
            edit_columns = ['id', 'task_display', 'owner', 'status', 'plan_start', 'plan_end', 'notes']
            column_names = {
                'id': 'ID', 'task_display': '任務名稱', 'owner': '負責單位', 'status': '狀態',
                'plan_start': '計劃開始', 'plan_end': '計劃完成', 'notes': '備註'
            }

        # 可編輯的任務表格
        edited_tasks_df = st.data_editor(
            display_tasks[edit_columns].rename(columns=column_names),
            column_config={
                "ID": st.column_config.NumberColumn("ID", disabled=True, width="small"),
                "任務名稱": st.column_config.TextColumn("任務名稱", width="large"),
                "負責單位": st.column_config.SelectboxColumn("負責單位", options=owner_options, width="medium"),
                "狀態": st.column_config.SelectboxColumn("狀態", options=["Done", "Going", "Delay"], width="small"),
                "計劃開始": st.column_config.DateColumn("計劃開始", format="YYYY-MM-DD"),
                "計劃完成": st.column_config.DateColumn("計劃完成", format="YYYY-MM-DD"),
                "計劃天數": st.column_config.NumberColumn("計劃天數", width="small", disabled=True),
                "實際開始": st.column_config.DateColumn("實際開始", format="YYYY-MM-DD"),
                "實際完成": st.column_config.DateColumn("實際完成", format="YYYY-MM-DD"),
                "完成%": st.column_config.NumberColumn("完成%", min_value=0, max_value=100, format="%.0f%%", width="small"),
                "誤差天數": st.column_config.NumberColumn("誤差天數", width="small", disabled=True),
                "備註": st.column_config.TextColumn("備註", width="large"),
            },
            num_rows="dynamic",  # 允許新增/刪除行
            use_container_width=True,
            hide_index=True,
            key="task_editor"
        )

        # 儲存變更
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            if st.button("💾 儲存所有變更", type="primary", use_container_width=True):
                # 還原欄位名稱
                reverse_column_names = {v: k for k, v in column_names.items()}
                edited_tasks_df_copy = edited_tasks_df.rename(columns=reverse_column_names)

                # 清理任務名稱（移除層級標記）
                if 'task_display' in edited_tasks_df_copy.columns:
                    def clean_task_name(x):
                        if pd.isna(x):
                            return ''
                        # 移除所有層級符號
                        cleaned = str(x)
                        cleaned = cleaned.replace('■ ', '')  # 主項目
                        cleaned = cleaned.replace('├─ ', '')  # 次項目
                        cleaned = cleaned.replace('└─ ', '')  # 次次項目
                        cleaned = cleaned.strip()
                        return cleaned

                    edited_tasks_df_copy['task'] = edited_tasks_df_copy['task_display'].apply(clean_task_name)
                    edited_tasks_df_copy = edited_tasks_df_copy.drop(columns=['task_display'])

                # ========== 資料驗證 ==========
                validation_errors = []

                for idx, row in edited_tasks_df_copy.iterrows():
                    task_id = idx + 1

                    # 1. 必填欄位檢查
                    if pd.isna(row.get('task')) or str(row.get('task', '')).strip() == '':
                        validation_errors.append(f"第 {task_id} 行：任務名稱不能為空")

                    if pd.isna(row.get('owner')) or str(row.get('owner', '')).strip() == '':
                        validation_errors.append(f"第 {task_id} 行：負責單位不能為空")

                    if pd.isna(row.get('status')) or str(row.get('status', '')).strip() == '':
                        validation_errors.append(f"第 {task_id} 行：狀態不能為空")

                    # 2. 日期邏輯檢查
                    plan_start = row.get('plan_start')
                    plan_end = row.get('plan_end')

                    if pd.notna(plan_start) and pd.notna(plan_end):
                        if pd.to_datetime(plan_start) > pd.to_datetime(plan_end):
                            validation_errors.append(f"第 {task_id} 行：計劃開始日期 ({plan_start}) 不能晚於計劃完成日期 ({plan_end})")

                    # 檢查實際日期
                    if 'actual_start' in row and 'actual_end' in row:
                        actual_start = row.get('actual_start')
                        actual_end = row.get('actual_end')

                        if pd.notna(actual_start) and pd.notna(actual_end):
                            if pd.to_datetime(actual_start) > pd.to_datetime(actual_end):
                                validation_errors.append(f"第 {task_id} 行：實際開始日期不能晚於實際完成日期")

                    # 3. 百分比範圍檢查
                    if 'progress_pct' in row:
                        progress = row.get('progress_pct')
                        if pd.notna(progress):
                            try:
                                progress_val = float(progress)
                                if progress_val < 0 or progress_val > 100:
                                    validation_errors.append(f"第 {task_id} 行：完成百分比必須在 0-100 之間（目前：{progress_val}）")
                            except (ValueError, TypeError):
                                validation_errors.append(f"第 {task_id} 行：完成百分比格式錯誤")

                # 顯示驗證錯誤
                if validation_errors:
                    st.error("❌ 資料驗證失敗，請修正以下錯誤：")
                    for error in validation_errors[:10]:  # 最多顯示 10 個錯誤
                        st.error(f"• {error}")
                    if len(validation_errors) > 10:
                        st.error(f"... 還有 {len(validation_errors) - 10} 個錯誤未顯示")
                else:
                    # 驗證通過，儲存資料
                    # 儲存到歷史記錄（用於撤銷/重做）
                    if len(st.session_state['edit_history']) == 0 or not st.session_state['edited_all_tasks'].equals(st.session_state['edit_history'][-1]):
                        # 清除重做歷史
                        st.session_state['edit_history'] = st.session_state['edit_history'][:st.session_state['history_index'] + 1]
                        # 加入新歷史
                        st.session_state['edit_history'].append(st.session_state['edited_all_tasks'].copy())
                        st.session_state['history_index'] = len(st.session_state['edit_history']) - 1
                        # 限制歷史記錄數量（最多 20 步）
                        if len(st.session_state['edit_history']) > 20:
                            st.session_state['edit_history'] = st.session_state['edit_history'][-20:]
                            st.session_state['history_index'] = 19

                    # 更新 edited_all_tasks 的對應欄位
                    for col in edit_columns:
                        if col in edited_tasks_df_copy.columns:
                            st.session_state['edited_all_tasks'][col] = edited_tasks_df_copy[col]

                    # 重新計算 ID
                    st.session_state['edited_all_tasks']['id'] = range(1, len(st.session_state['edited_all_tasks']) + 1)

                    # 更新時間戳記
                    st.session_state['last_edit_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                    # 儲存當前狀態到歷史
                    st.session_state['edit_history'].append(st.session_state['edited_all_tasks'].copy())
                    st.session_state['history_index'] = len(st.session_state['edit_history']) - 1

                    st.success(f"✅ 已儲存 {len(edited_tasks_df_copy)} 個任務的變更｜所有圖表已同步")
                    st.info("💡 所有分頁的圖表已更新，前往「匯出」分頁下載 Excel")
                    st.rerun()

        with col2:
            if st.button("🔄 重置為原始資料", use_container_width=True):
                st.session_state['edited_project_info'] = data['project_info'].copy()
                st.session_state['edited_all_tasks'] = data['tasks'].copy()
                st.session_state['edited_system_tasks'] = data['system_tasks'].copy()
                if 'last_edit_time' in st.session_state:
                    del st.session_state['last_edit_time']
                st.success("✅ 已重置為原始資料")
                st.rerun()

        with col3:
            st.markdown("**提示：** 可直接在表格中編輯、新增或刪除行（點擊行號旁的 ✖️）")

        st.divider()
        st.divider()

        # ========== 系統時程編輯 ==========
        st.markdown("### 🏭 系統時程編輯")
        st.info("💡 編輯各區域的完成百分比、目標日期等資訊")

        # 系統時程編輯器
        if not df_system.empty and 'edited_system_tasks' in st.session_state:
            system_col1, system_col2 = st.columns([3, 1])

            with system_col1:
                # 只顯示區域（is_area == True）的項目
                area_tasks = st.session_state['edited_system_tasks'][
                    st.session_state['edited_system_tasks']['is_area'] == True
                ].copy()

                if not area_tasks.empty:
                    # 可編輯的系統時程表格
                    system_edit_columns = ['item', 'completion_pct', 'target_date', 'notes']
                    system_column_names = {
                        'item': '區域', 'completion_pct': '完成百分比',
                        'target_date': '目標日期', 'notes': '備註'
                    }

                    # 準備編輯用的數據 - 確保數據類型正確
                    area_tasks_for_edit = area_tasks[system_edit_columns].copy()

                    # 確保 completion_pct 是 float 類型
                    if 'completion_pct' in area_tasks_for_edit.columns:
                        area_tasks_for_edit['completion_pct'] = pd.to_numeric(
                            area_tasks_for_edit['completion_pct'],
                            errors='coerce'
                        ).fillna(0.0)

                    # 確保 target_date 是 datetime 類型（可以為 None）
                    if 'target_date' in area_tasks_for_edit.columns:
                        # 嘗試轉換為 datetime，失敗則設為 None
                        try:
                            area_tasks_for_edit['target_date'] = pd.to_datetime(
                                area_tasks_for_edit['target_date'],
                                errors='coerce'
                            )
                        except:
                            area_tasks_for_edit['target_date'] = None

                    # 確保 notes 是字符串類型
                    if 'notes' in area_tasks_for_edit.columns:
                        area_tasks_for_edit['notes'] = area_tasks_for_edit['notes'].fillna('').astype(str)

                    edited_system_df = st.data_editor(
                        area_tasks_for_edit.rename(columns=system_column_names),
                        column_config={
                            "區域": st.column_config.TextColumn("區域", disabled=True, width="medium"),
                            "完成百分比": st.column_config.NumberColumn(
                                "完成百分比",
                                min_value=0,
                                max_value=100,
                                format="%.1f%%",
                                width="small",
                                help="輸入 0-100 之間的數值（例如：75 代表 75%）"
                            ),
                            "目標日期": st.column_config.DateColumn("目標日期", format="YYYY-MM-DD"),
                            "備註": st.column_config.TextColumn("備註", width="large"),
                        },
                        use_container_width=True,
                        hide_index=True,
                        key="system_editor"
                    )

                    st.caption(f"📊 共有 {len(edited_system_df)} 個區域")

            with system_col2:
                st.markdown("**系統時程操作：**")

                if st.button("💾 儲存系統時程", type="primary", use_container_width=True):
                    # 還原欄位名稱
                    reverse_system_names = {v: k for k, v in system_column_names.items()}
                    edited_system_copy = edited_system_df.rename(columns=reverse_system_names)

                    # 更新 session_state 中的系統時程資料（只更新區域項目）
                    area_indices = area_tasks.index
                    for col in system_edit_columns:
                        if col in edited_system_copy.columns:
                            st.session_state['edited_system_tasks'].loc[area_indices, col] = edited_system_copy[col].values

                    st.session_state['last_edit_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    st.success("✅ 系統時程已更新")
                    st.rerun()

                if st.button("🔄 重置系統時程", use_container_width=True):
                    st.session_state['edited_system_tasks'] = data['system_tasks'].copy()
                    st.success("✅ 已重置為原始系統時程")
                    st.rerun()
        else:
            st.warning("⚠️ 未偵測到系統時程資料")

    # Tab 7: 週報生成
    with tab7:
        st.subheader("📝 專案週報生成")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            report_content = generate_weekly_report(data, datetime.combine(report_date, datetime.min.time()))
            st.markdown(report_content)
        
        with col2:
            st.markdown("### 📥 下載報表")
            
            st.download_button(
                label="📄 下載 Markdown",
                data=report_content,
                file_name=f"週報_{report_date.strftime('%Y%m%d')}.md",
                mime="text/markdown",
            )
            
            st.divider()
            
            st.markdown("### 📊 快速統計")
            summary = generate_status_summary(data)
            
            st.metric("完成率", f"{summary['done']/summary['total']*100:.1f}%")
            st.metric("延遲項目", summary['delay'])
            st.metric("本週到期", len(summary['upcoming']))

            st.divider()

            # 通知功能
            st.markdown("### 📢 發送通知")

            if NOTIFICATIONS_AVAILABLE:
                # 初始化通知配置 session_state
                if 'notification_config' not in st.session_state:
                    st.session_state['notification_config'] = {
                        'teams_enabled': False,
                        'teams_webhook': '',
                        'email_enabled': False,
                        'email_recipients': '',
                    }

                with st.expander("⚙️ 通知設定", expanded=False):
                    # Teams 設定
                    teams_enabled = st.checkbox("啟用 Teams 通知", value=st.session_state['notification_config']['teams_enabled'])
                    teams_webhook = st.text_input(
                        "Teams Webhook URL",
                        value=st.session_state['notification_config']['teams_webhook'],
                        type="password",
                        help="請輸入 Microsoft Teams Incoming Webhook URL"
                    )

                    # Email 設定
                    email_enabled = st.checkbox("啟用 Email 通知", value=st.session_state['notification_config']['email_enabled'])
                    email_recipients = st.text_input(
                        "Email 收件人 (逗號分隔)",
                        value=st.session_state['notification_config']['email_recipients'],
                        help="例如：user1@company.com,user2@company.com"
                    )

                    if st.button("💾 儲存通知設定"):
                        st.session_state['notification_config'] = {
                            'teams_enabled': teams_enabled,
                            'teams_webhook': teams_webhook,
                            'email_enabled': email_enabled,
                            'email_recipients': email_recipients,
                        }
                        st.success("✅ 通知設定已儲存")

                # 發送通知按鈕
                notify_col1, notify_col2 = st.columns(2)

                with notify_col1:
                    if st.button("📊 發送週報", use_container_width=True):
                        config = NotificationConfig()
                        config.teams_enabled = st.session_state['notification_config']['teams_enabled']
                        config.teams_webhook_url = st.session_state['notification_config']['teams_webhook']

                        notifier = ProjectNotifier(config)
                        notifier.send_weekly_report(report_content, project_info.get('project_name', 'OHTC 專案'))
                        st.success("✅ 週報已發送！")

                with notify_col2:
                    if st.button("⚠️ 發送延遲警報", use_container_width=True):
                        delay_tasks = df_tasks[df_tasks['status'] == 'Delay'].to_dict('records')
                        if delay_tasks:
                            config = NotificationConfig()
                            config.teams_enabled = st.session_state['notification_config']['teams_enabled']
                            config.teams_webhook_url = st.session_state['notification_config']['teams_webhook']

                            notifier = ProjectNotifier(config)
                            notifier.send_delay_alert(delay_tasks, project_info.get('project_name', 'OHTC 專案'))
                            st.success(f"✅ 已發送 {len(delay_tasks)} 個延遲項目的警報！")
                        else:
                            st.info("💡 目前沒有延遲項目")

                if st.button("📈 發送每日摘要", use_container_width=True):
                    config = NotificationConfig()
                    config.teams_enabled = st.session_state['notification_config']['teams_enabled']
                    config.teams_webhook_url = st.session_state['notification_config']['teams_webhook']

                    notifier = ProjectNotifier(config)
                    notifier.send_daily_summary(summary, project_info.get('project_name', 'OHTC 專案'))
                    st.success("✅ 每日摘要已發送！")
            else:
                st.warning("⚠️ 通知功能不可用：notifications.py 模組未找到")

    # Tab 8: 匯出
    with tab8:
        st.subheader("⬇️ 匯出資料")

        # 檢查是否有編輯過的資料
        has_edits = 'edited_all_tasks' in st.session_state or 'edited_project_info' in st.session_state

        if has_edits:
            st.info("💡 偵測到您在「專案編輯」分頁有進行修改，匯出將使用最新的編輯資料")

        # 顯示額外分頁資料
        st.markdown("### 📋 額外分頁資料預覽")

        extra_tabs = []
        if not data.get('progress_stats', pd.DataFrame()).empty:
            extra_tabs.append("進度統計")
        if not data.get('eq_list', pd.DataFrame()).empty:
            extra_tabs.append("EQ 工作清單")
        if data.get('layout_images') and len(data.get('layout_images', [])) > 0:
            extra_tabs.append("Layout 圖片")

        if extra_tabs:
            extra_tab_objects = st.tabs(extra_tabs)

            tab_idx = 0
            if not data.get('progress_stats', pd.DataFrame()).empty:
                with extra_tab_objects[tab_idx]:
                    st.markdown("#### 📊 進度統計")
                    df_stats = data['progress_stats']

                    # 顯示統計摘要
                    if not df_stats.empty:
                        # 計算各項目完成數
                        items = ['C鋼', '軌道', 'HID', '圖資', 'OHB', 'CycleTest']
                        summary_cols = st.columns(len(items))
                        for idx, item in enumerate(items):
                            target_col = f'{item}_目標'
                            actual_col = f'{item}_實際'
                            if target_col in df_stats.columns and actual_col in df_stats.columns:
                                total = df_stats[target_col].notna().sum()
                                done = df_stats[actual_col].notna().sum()
                                with summary_cols[idx]:
                                    st.metric(item, f"{done}/{total}")

                        st.divider()

                    # 顯示詳細表格
                    st.dataframe(df_stats, use_container_width=True, height=400)
                tab_idx += 1

            if not data.get('eq_list', pd.DataFrame()).empty:
                with extra_tab_objects[tab_idx]:
                    st.markdown("#### 🔧 EQ 工作清單")
                    st.dataframe(data['eq_list'], use_container_width=True, height=400)
                tab_idx += 1

            if data.get('layout_images') and len(data.get('layout_images', [])) > 0:
                with extra_tab_objects[tab_idx]:
                    st.markdown("#### 🖼️ Layout 圖片")
                    layout_images = data.get('layout_images', [])
                    st.write(f"共找到 {len(layout_images)} 張圖片")
                    for idx, img_bytes in enumerate(layout_images):
                        try:
                            st.image(img_bytes, caption=f"Layout 圖片 {idx + 1}", use_container_width=True)
                        except Exception as e:
                            st.error(f"無法顯示圖片 {idx + 1}: {str(e)}")
        else:
            st.info("📝 此檔案中沒有「進度統計」、「EQ 工作清單」或「Layout 圖片」分頁")

        st.divider()
        st.markdown("### 💾 下載檔案")

        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown("### 📊 Excel 完整匯出")
            st.write("保持原始格式，匯出更新後的排程表")

            if st.button("🔄 生成 Excel", type="primary"):
                try:
                    # 優先使用編輯過的資料
                    tasks_to_export = st.session_state.get('edited_all_tasks', df_tasks)
                    project_to_export = st.session_state.get('edited_project_info', project_info)

                    # 建立包含編輯資料的 data 字典
                    export_data = {
                        'project_info': project_to_export,
                        'tasks': tasks_to_export,
                        'system_tasks': data.get('system_tasks'),
                        'raw_software': data.get('raw_software'),
                    }

                    excel_output = export_updated_excel(export_data, uploaded_file, tasks_to_export)

                    # 生成檔案名稱：專案名稱+安裝排程表+_日期+_v版號
                    export_filename = generate_export_filename(
                        uploaded_file.name,
                        project_to_export.get('project_name', 'OHTC')
                    )

                    st.download_button(
                        label="⬇️ 下載 Excel",
                        data=excel_output,
                        file_name=export_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.success(f"✅ Excel 已生成：{export_filename}")
                except Exception as e:
                    st.error(f"匯出失敗: {str(e)}")
                    st.exception(e)

        with col2:
            st.markdown("### 📋 CSV 匯出")
            st.write("任務清單輕量匯出")

            # 使用編輯過的資料
            csv_data = st.session_state.get('edited_all_tasks', df_tasks)
            csv = csv_data.to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                label="⬇️ 下載 CSV",
                data=csv,
                file_name=f"任務清單_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv"
            )

        with col3:
            st.markdown("### 📈 JSON 匯出")
            st.write("結構化資料匯出，適合程式處理")

            # 使用編輯過的資料
            json_project = st.session_state.get('edited_project_info', project_info)
            json_tasks = st.session_state.get('edited_all_tasks', df_tasks)

            json_data = {
                'project_info': json_project,
                'task_count': len(json_tasks),
                'exported_at': datetime.now().isoformat(),
            }

            st.download_button(
                label="⬇️ 下載 JSON",
                data=json.dumps(json_data, ensure_ascii=False, indent=2, default=str),
                file_name=f"專案摘要_{datetime.now().strftime('%Y%m%d')}.json",
                mime="application/json"
            )


if __name__ == "__main__":
    main()
