"""
OHTC 專案排程表模板生成器
========================
用於建立新專案的標準化排程表 Excel 模板
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference
from datetime import datetime, timedelta
import argparse


class ScheduleTemplateGenerator:
    """排程表模板生成器"""
    
    # 樣式定義
    COLORS = {
        'header_bg': 'FF4472C4',      # 藍色
        'header_font': 'FFFFFFFF',    # 白色
        'done_bg': 'FF92D050',        # 綠色
        'going_bg': 'FFFFEB9C',       # 黃色
        'delay_bg': 'FFFF6B6B',       # 紅色
        'border': 'FFB4B4B4',         # 灰色邊框
    }
    
    def __init__(self):
        self.wb = Workbook()
        self._setup_styles()
    
    def _setup_styles(self):
        """設定樣式"""
        self.header_font = Font(bold=True, color=self.COLORS['header_font'], size=11)
        self.header_fill = PatternFill(start_color=self.COLORS['header_bg'], 
                                       end_color=self.COLORS['header_bg'], 
                                       fill_type='solid')
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        self.thin_border = Border(
            left=Side(style='thin', color=self.COLORS['border']),
            right=Side(style='thin', color=self.COLORS['border']),
            top=Side(style='thin', color=self.COLORS['border']),
            bottom=Side(style='thin', color=self.COLORS['border'])
        )
    
    def create_software_schedule(self, project_info: dict, tasks: list = None):
        """建立軟體時程表"""
        ws = self.wb.active
        ws.title = '軟體時程'
        
        # 設定欄寬
        column_widths = {
            'A': 35,  # 項目
            'B': 5,   # 空白
            'C': 15,  # 負責單位
            'D': 8,   # 實際完成進度
            'E': 10,  # 實際完成百分比
            'F': 10,  # 預計完成百分比
            'G': 8,   # 剩餘天數
            'H': 8,   # 進度
            'I': 12,  # 計劃開始日期
            'J': 12,  # 計劃完成日期
            'K': 8,   # 計劃天數
            'L': 12,  # 實際開始日期
            'M': 12,  # 實際完成日期
            'N': 8,   # 實際天數
            'O': 8,   # 誤差天數
            'P': 10,  # 協調時間
            'Q': 10,  # 協調人力
            'R': 10,  # 協調區域
            'S': 10,  # 協調設備
            'T': 20,  # 備註
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 專案標題
        ws['A1'] = f"{project_info.get('name', 'OHTC專案')}_排程表"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # 專案資訊區
        info_rows = [
            ('專案工令', 'project_code'),
            ('專案名稱', 'name'),
            ('專案負責', 'lead'),
        ]
        
        for i, (label, key) in enumerate(info_rows, start=3):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = '請輸入' + label
            ws[f'C{i}'] = project_info.get(key, '')
        
        # 計畫開始日
        ws['I4'] = '計畫開始日'
        ws['J4'] = project_info.get('start_date', datetime.now())
        
        # 標題行
        headers = [
            '項目', '', '負責單位', '實際完成\n進度', '實際完成\n百分比', 
            '預計完成\n百分比', '剩餘\n天數', '進度', '計劃開始日期', '計劃完成日期',
            '計劃\n天數', '實際開始日期', '實際完成日期', '實際\n天數', '誤差\n天數',
            '協調時間', '協調人力', '協調區域', '協調設備', '備註'
        ]
        
        header_row = 6
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        
        # 凍結窗格
        ws.freeze_panes = 'A7'
        
        # 新增預設任務
        default_tasks = tasks or [
            {'task': 'Server安裝及設定', 'owner': 'IT', 'days': 5},
            {'task': '  Server安裝', 'owner': 'IT', 'days': 3},
            {'task': '  OS安裝', 'owner': 'IT', 'days': 1},
            {'task': '  遠端連線環境建立', 'owner': 'IT', 'days': 1},
            {'task': 'OHTC軟體安裝及設定', 'owner': 'OHTC', 'days': 7},
            {'task': '  HA設定及服務建立', 'owner': 'OHTC', 'days': 2},
            {'task': '  OHTC1設定', 'owner': 'OHTC', 'days': 1},
            {'task': '  OHTC2設定', 'owner': 'OHTC', 'days': 1},
            {'task': '現場主機狀態確認', 'owner': 'OHTC', 'days': 2},
            {'task': 'Golden驗證', 'owner': 'OHTL', 'days': 10},
        ]
        
        start_date = project_info.get('start_date', datetime.now())
        current_date = start_date
        
        for i, task in enumerate(default_tasks, start=7):
            ws.cell(row=i, column=1, value=task['task']).border = self.thin_border
            ws.cell(row=i, column=3, value=task['owner']).border = self.thin_border
            
            # 計劃日期
            end_date = current_date + timedelta(days=task['days'])
            ws.cell(row=i, column=9, value=current_date).border = self.thin_border
            ws.cell(row=i, column=10, value=end_date).border = self.thin_border
            ws.cell(row=i, column=11, value=task['days']).border = self.thin_border
            
            # 進度欄位
            ws.cell(row=i, column=8, value='').border = self.thin_border
            
            # 設定日期格式
            ws.cell(row=i, column=9).number_format = 'YYYY-MM-DD'
            ws.cell(row=i, column=10).number_format = 'YYYY-MM-DD'
            
            current_date = end_date
        
        # 新增條件格式（進度狀態顏色）
        done_fill = PatternFill(start_color=self.COLORS['done_bg'], 
                                end_color=self.COLORS['done_bg'], 
                                fill_type='solid')
        going_fill = PatternFill(start_color=self.COLORS['going_bg'], 
                                 end_color=self.COLORS['going_bg'], 
                                 fill_type='solid')
        delay_fill = PatternFill(start_color=self.COLORS['delay_bg'], 
                                 end_color=self.COLORS['delay_bg'], 
                                 fill_type='solid')
        
        ws.conditional_formatting.add('H7:H100',
            FormulaRule(formula=['$H7="Done"'], fill=done_fill))
        ws.conditional_formatting.add('H7:H100',
            FormulaRule(formula=['$H7="Going"'], fill=going_fill))
        ws.conditional_formatting.add('H7:H100',
            FormulaRule(formula=['$H7="Delay"'], fill=delay_fill))
        
        # 新增資料驗證（下拉選單）
        from openpyxl.worksheet.datavalidation import DataValidation
        status_dv = DataValidation(type="list", formula1='"Done,Going,Delay"', allow_blank=True)
        status_dv.error = '請選擇有效的狀態'
        status_dv.errorTitle = '無效輸入'
        ws.add_data_validation(status_dv)
        status_dv.add('H7:H100')
        
        return ws
    
    def create_system_schedule(self):
        """建立系統時程表"""
        ws = self.wb.create_sheet('系統時程_C')
        
        # 設定欄寬
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        
        # 標題
        headers = ['項目', '計劃完成日期', '完成\n百分比', '備註']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
        
        # 預設區域
        areas = ['區域A', '區域B', '區域C', '區域D', '區域E', '區域F', '區域G', '區域H']
        sub_items = ['走行', '  踩點', '  提速', '安全', '  AREA SENSOR驗證', 
                     '取放', '  OHB Teaching', '  EQ Teaching', '系統', '  MCS測試']
        
        row = 7
        for area in areas:
            ws.cell(row=row, column=1, value=area)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=3, value=0)
            row += 1
            
            for item in sub_items:
                ws.cell(row=row, column=1, value=item)
                ws.cell(row=row, column=3, value=0)
                row += 1
        
        # 新增完成百分比色階
        ws.conditional_formatting.add(f'C7:C{row}',
            ColorScaleRule(start_type='num', start_value=0, start_color='FFFF6B6B',
                          mid_type='num', mid_value=0.5, mid_color='FFFFEB9C',
                          end_type='num', end_value=1, end_color='FF92D050'))
        
        return ws
    
    def create_engineering_progress(self):
        """建立工程進度確認表"""
        ws = self.wb.create_sheet('工程_工作進度確認表')
        
        # 標題行
        headers = ['', '區域', '項目', 'C鋼', '軌道', '', 'HID', '', '圖資', '', 
                   'OHB', '', 'Cycle Test', '', 'EQ Teaching', 'Hot Run', 'RTD Test', 'Release']
        sub_headers = ['', '', '', '', '目標', '實際', '目標', '實際', '目標', '實際',
                      '目標', '實際', '目標', '實際', '', '', '', '']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
        
        for col, header in enumerate(sub_headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.alignment = self.center_align
        
        # 預設 Bay
        bays = [('A', 'Bay 1'), ('H', 'Bay 2'), ('H', 'Bay 3'), ('H', 'Bay 4'),
                ('H', 'Bay 5'), ('H', 'Bay 6'), ('G', 'Bay 7'), ('G', 'Bay 8'),
                ('G', 'Bay 9'), ('G', 'Bay 10'), ('F', 'Bay 11'), ('F', 'Bay 12')]
        
        for i, (area, bay) in enumerate(bays, start=4):
            ws.cell(row=i, column=2, value=area)
            ws.cell(row=i, column=3, value=bay)
        
        return ws
    
    def create_eq_list(self):
        """建立 EQ 工作清單"""
        ws = self.wb.create_sheet('EQ 工作清單')
        
        headers = ['區域', 'EQ', 'Port ID', '安裝', '校點', 'PIO測試', 
                   'OHTC取放', 'MCS取放', 'MES取放']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        ws.cell(row=1, column=1, value='尚未有資料')
        
        return ws
    
    def create_location_map(self):
        """建立機台分佈位置表"""
        ws = self.wb.create_sheet('機台分佈位置')
        ws.cell(row=1, column=1, value='尚未有資料')
        return ws
    
    def create_fab_map(self):
        """建立 Fab 測試分區圖示"""
        ws = self.wb.create_sheet('Fab測試分區圖示')
        
        headers = ['區域', 'OHB數量']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        areas = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        for i, area in enumerate(areas, start=2):
            ws.cell(row=i, column=1, value=area)
        
        return ws
    
    def generate(self, project_info: dict, output_path: str = None):
        """生成完整模板"""
        
        # 建立所有工作表
        self.create_software_schedule(project_info)
        self.create_system_schedule()
        self.create_engineering_progress()
        self.create_eq_list()
        self.create_location_map()
        self.create_fab_map()
        
        # 儲存
        if output_path is None:
            project_name = project_info.get('name', 'OHTC專案').replace(' ', '_')
            date_str = datetime.now().strftime('%Y%m%d')
            output_path = f"{project_name}_排程表_{date_str}.xlsx"
        
        self.wb.save(output_path)
        print(f"✅ 模板已生成: {output_path}")
        
        return output_path


def main():
    parser = argparse.ArgumentParser(description='OHTC 專案排程表模板生成器')
    parser.add_argument('-n', '--name', default='新專案', help='專案名稱')
    parser.add_argument('-c', '--code', default='', help='專案工令')
    parser.add_argument('-l', '--lead', default='', help='專案負責人')
    parser.add_argument('-s', '--start', default=None, help='開始日期 (YYYY-MM-DD)')
    parser.add_argument('-o', '--output', default=None, help='輸出檔案路徑')
    
    args = parser.parse_args()
    
    project_info = {
        'name': args.name,
        'project_code': args.code,
        'lead': args.lead,
        'start_date': datetime.strptime(args.start, '%Y-%m-%d') if args.start else datetime.now(),
    }
    
    generator = ScheduleTemplateGenerator()
    generator.generate(project_info, args.output)


if __name__ == '__main__':
    main()
